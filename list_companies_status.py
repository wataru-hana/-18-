#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""実装済みと未実装の企業をリストアップ"""

import yaml
from openpyxl import load_workbook

# sites.yamlから企業リストを取得
with open('config/sites.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)
    sites = config.get('sites', [])

# Excelファイルから実装済み企業を取得
excel_path = 'price_results_v2_20251104_220253.xlsx'
wb = load_workbook(excel_path)

# 最新のシートを探す
target_sheet = None
for sheet_name in reversed(wb.sheetnames):
    if sheet_name.startswith('価格情報_'):
        target_sheet = wb[sheet_name]
        break

implemented_companies = set()
if target_sheet:
    for row_idx in range(2, target_sheet.max_row + 1):
        company_name = target_sheet.cell(row=row_idx, column=1).value
        if company_name:
            implemented_companies.add(str(company_name))

# 企業を分類
all_companies = {}
for site in sites:
    name = site.get('name', '')
    if name:
        all_companies[name] = {
            'region': site.get('region', ''),
            'url': site.get('price_url', ''),
            'extractor_type': site.get('extractor_type', 'auto'),
            'implemented': name in implemented_companies
        }

# 実装済みと未実装に分類
implemented = []
not_implemented = []

for name, info in sorted(all_companies.items()):
    if info['implemented']:
        implemented.append((name, info))
    else:
        not_implemented.append((name, info))

# 結果を表示
print("="*80)
print("実装状況一覧")
print("="*80)
print(f"\n総企業数: {len(all_companies)}社")
print(f"実装済み: {len(implemented)}社")
print(f"未実装: {len(not_implemented)}社")
print("\n" + "="*80)

print("\n【実装済み企業】")
print("="*80)
for i, (name, info) in enumerate(implemented, 1):
    print(f"{i}. {name}")
    print(f"   地域: {info['region']}")
    print(f"   抽出方法: {info['extractor_type']}")
    print()

print("\n【未実装企業】")
print("="*80)
for i, (name, info) in enumerate(not_implemented, 1):
    print(f"{i}. {name}")
    print(f"   地域: {info['region']}")
    print(f"   抽出方法: {info['extractor_type']}")
    if info['url']:
        print(f"   URL: {info['url']}")
    print()










