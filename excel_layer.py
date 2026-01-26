#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel層：STDテーブルをExcelの「価格一覧表」シートに直接書き込む
- 価格は数値型（int/float）のみ
- normalize_key()を使ったヘッダー照合
- company_id優先の会社行マッチング
- 書き込み前にセルクリア
- fullCalcOnLoadをON
"""

import yaml
import logging
import re
from pathlib import Path
from typing import Dict, Optional, Tuple, List
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties
from data_structures import STDTable, STANDARD_ITEMS

logger = logging.getLogger(__name__)


def normalize_key(key: str) -> str:
    """キーを正規化（ExcelヘッダーとSTDキーの照合用）
    
    仕様：
    - 全角スペース、半角スペースを削除
    - 全角％→半角%に統一
    - 「_」「-」などの記号は除去
    - 大文字小文字は無視
    - 同義語aliasを適用（正規化後）
    """
    if not key:
        return ''
    
    # 文字列に変換
    key = str(key)
    
    # 全角％→半角%に統一
    key = key.replace('％', '%')
    
    # 全角スペース、半角スペースを削除
    key = key.replace(' ', '').replace('　', '')
    
    # 記号を除去（「_」「-」など）
    key = re.sub(r'[_\-\－＿]', '', key)
    
    # 大文字小文字を無視（小文字に統一）
    key = key.lower()
    
    # 同義語aliasを適用（正規化後）
    # アルミ缶関連のチェック（「アルミ缶」を含むキーはすべて「アルミ缶」に正規化）
    if 'アルミ缶' in key or key in ['ubc', '缶', 'バラ缶', 'プレス缶']:
        key = 'アルミ缶'
    # ピカ銅関連
    elif key in ['ピカ銅', 'ピカ線', '上銅']:
        key = 'ピカ銅'
    # 雑線80%関連
    elif '雑線80%' in key or key in ['一本線80%', '上線80%']:
        key = '雑線80%'
    # 雑線60%-65%関連
    elif '雑線60%-65%' in key or '雑60%-65%' in key or key in ['三本線60%', '上線60%']:
        key = '雑線60%-65%'
    # ステンレス304関連
    elif 'ステンレス304' in key or key in ['ステンレス', 'sus304', '304']:
        key = 'ステンレス304'
    # 鉛バッテリー関連
    elif '鉛バッテリー' in key or key in ['バッテリー', '廃バッテリー']:
        key = '鉛バッテリー'
    
    return key


def normalize_header_name(header: str) -> str:
    """ヘッダー名を正規化（Excelヘッダー行から列位置を確定するため）
    
    全角空白/半角空白除去、改行除去、連続空白圧縮
    """
    if not header:
        return ''
    
    header = str(header)
    
    # 改行を除去
    header = header.replace('\n', '').replace('\r', '')
    
    # 全角スペース、半角スペースを削除
    header = header.replace(' ', '').replace('　', '')
    
    # 連続空白を圧縮（既に空白を削除したので不要だが、念のため）
    header = re.sub(r'\s+', '', header)
    
    return header.strip()


def normalize_company_name(name: str) -> str:
    """企業名を正規化（括弧や全角/半角、スペース差分を吸収）
    
    Args:
        name: 企業名
        
    Returns:
        正規化された企業名
    """
    if not name:
        return ''
    
    name = str(name).strip()
    
    # 括弧の種類を統一（全角括弧に統一）
    name = name.replace('(', '（').replace(')', '）')
    
    # 全角スペースと半角スペースを統一（全角スペースに統一）
    name = name.replace(' ', '　')
    
    # 連続スペースを1つに
    name = re.sub(r'　+', '　', name)
    
    return name


def load_output_tables_config(config_path: str = 'config/output_tables.yaml') -> list:
    """
    output_tables.yamlを読み込む
    
    Args:
        config_path: output_tables.yamlのパス
        
    Returns:
        output_tablesのリスト
    """
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('output_tables', [])
    except Exception as e:
        logger.error(f"output_tables.yamlの読み込みエラー: {str(e)}")
        return []


def load_sites_config(config_path: str = 'config/sites.yaml') -> Dict[str, Dict[str, str]]:
    """
    sites.yamlからcompany_id、company_name、正規化後のcompany_nameのマッピングを読み込む
    
    Args:
        config_path: sites.yamlのパス
        
    Returns:
        {company_id: {'name': company_name, 'normalized_name': normalized_name}} の辞書
    """
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            sites = config.get('sites', [])
            result = {}
            for site in sites:
                company_id = site.get('id') or site.get('company_id') or site.get('name', '')
                company_name = site.get('name', '')
                if company_id and company_name:
                    normalized_name = normalize_company_name(company_name)
                    result[company_id] = {
                        'name': company_name,
                        'normalized_name': normalized_name
                    }
            return result
    except Exception as e:
        logger.warning(f"sites.yamlの読み込みエラー: {str(e)}")
        return {}


def find_sheet_in_workbook(wb, target_sheet_name: str) -> Optional[str]:
    """
    ワークブック内でシート名を探す（全角・半角両方に対応）
    
    Args:
        wb: openpyxlのワークブックオブジェクト
        target_sheet_name: 探すシート名
        
    Returns:
        実際のシート名（見つからない場合はNone）
    """
    for sheet_name in wb.sheetnames:
        # 完全一致または部分一致で探す
        if (target_sheet_name == sheet_name or 
            target_sheet_name in str(sheet_name) or 
            str(sheet_name) in target_sheet_name):
            return sheet_name
    return None


def extract_numeric_price(price) -> Optional[float]:
    """価格を数値型に変換（文字列の場合は数値のみを抽出）
    
    Args:
        price: 価格（int, float, str, Noneなど）
        
    Returns:
        数値（float）またはNone
    """
    if price is None:
        return None
    
    # 既に数値型の場合はそのまま
    if isinstance(price, (int, float)):
        return float(price)
    
    # 文字列の場合、数値を抽出
    if isinstance(price, str):
        # 数値を抽出（カンマや円マークを除去）
        price_match = re.search(r'(\d{1,4}(?:[,，]\d{3})*(?:\.\d+)?)', str(price))
        if price_match:
            price_value_str = price_match.group(1).replace(',', '').replace('，', '')
            try:
                return float(price_value_str)
            except ValueError:
                return None
    
    return None


def write_std_to_excel(
    std_table: STDTable,
    output_tables_config_path: str = 'config/output_tables.yaml',
    sites_config_path: str = 'config/sites.yaml'
) -> bool:
    """
    STDテーブルをExcelの「価格一覧表」シートに直接書き込む
    
    処理内容:
    - 価格は数値型（int/float）のみ書き込む
    - ヘッダー行から列位置を確定（normalize_key()を使用）
    - 会社行の特定はcompany_idを最優先、次にcompany_name正規化一致
    - 書き込み前にセルをクリア
    - fullCalcOnLoadをON
    
    Args:
        std_table: STDテーブル
        output_tables_config_path: output_tables.yamlのパス
        sites_config_path: sites.yamlのパス
        
    Returns:
        成功した場合True、失敗した場合False
    """
    output_tables = load_output_tables_config(output_tables_config_path)
    company_info = load_sites_config(sites_config_path)
    
    if not output_tables:
        logger.warning("output_tables.yamlに設定がありません")
        return False
    
    success_count = 0
    
    for table_config in output_tables:
        # enabledがFalseの場合はスキップ
        if not table_config.get('enabled', True):
            logger.info(f"スキップ（enabled=false）: {table_config.get('description', '')}")
            continue
        
        excel_file = table_config.get('excel_file', '')
        sheet_name = table_config.get('sheet_name', '')
        description = table_config.get('description', '')
        
        if not excel_file or not sheet_name:
            logger.warning(f"設定が不完全です: {table_config}")
            continue
        
        logger.info(f"Excelに転記中: {excel_file} - {sheet_name} ({description})")
        
        try:
            # Excelファイルを読み込む
            excel_path = Path(excel_file)
            if not excel_path.exists():
                logger.error(f"Excelファイルが見つかりません: {excel_file}")
                continue
            
            # 数式を保持するためにdata_only=Falseで読み込む（デフォルト）
            # 既存のシート（特にデバッグ情報シート）の数式を壊さないように注意
            wb = load_workbook(excel_file, data_only=False)
            
            # fullCalcOnLoadをON（開いた瞬間に再計算されるように）
            if getattr(wb, "calculation", None) is None:
                wb.calculation = CalcProperties()
            
            # 開いた瞬間に再計算させたい
            wb.calculation.fullCalcOnLoad = True
            
            # openpyxlのバージョン差異があるので、存在する属性だけ設定
            if hasattr(wb.calculation, "calcMode"):
                wb.calculation.calcMode = "auto"
            if hasattr(wb.calculation, "forceFullCalc"):
                wb.calculation.forceFullCalc = True
            
            # シートを探す
            actual_sheet_name = find_sheet_in_workbook(wb, sheet_name)
            if not actual_sheet_name:
                logger.error(f"シートが見つかりません: {sheet_name}")
                logger.info(f"利用可能なシート: {wb.sheetnames}")
                continue
            
            ws = wb[actual_sheet_name]
            logger.info(f"シート「{actual_sheet_name}」を読み込みました")
            
            # ヘッダー行（1行目、2列目以降）から材料名と列番号のマッピングを作成（normalize_key()を使用）
            header_to_col: Dict[str, int] = {}  # 正規化後のヘッダー名 -> 列番号
            header_original: Dict[str, str] = {}  # 正規化後のヘッダー名 -> 元のヘッダー名
            
            for col_idx in range(2, ws.max_column + 1):  # 2列目から（1列目は企業名）
                cell = ws.cell(row=1, column=col_idx)
                if cell.value:
                    header_name = normalize_header_name(str(cell.value))
                    if header_name:
                        normalized_header = normalize_key(header_name)
                        if normalized_header:
                            header_to_col[normalized_header] = col_idx
                            header_original[normalized_header] = str(cell.value).strip()
            
            logger.info(f"ヘッダー材料（正規化後）: {list(header_to_col.keys())[:10]}...")  # 最初の10個だけ
            
            # 企業名列（1列目）から企業名と行番号のマッピングを作成
            company_rows: Dict[str, int] = {}  # 正規化後の企業名 -> 行番号
            company_original: Dict[str, str] = {}  # 正規化後の企業名 -> 元の企業名
            
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                if cell.value:
                    company_name = str(cell.value).strip()
                    normalized_company_name = normalize_company_name(company_name)
                    if normalized_company_name:
                        company_rows[normalized_company_name] = row_idx
                        company_original[normalized_company_name] = company_name
            
            logger.info(f"既存の企業: {len(company_rows)}社")
            
            # STDテーブルをExcelに転記
            total_filled_count = 0
            total_companies = 0
            
            for company_id, items in std_table.items():
                total_companies += 1
                row_idx: Optional[int] = None
                matched_by: Optional[str] = None
                
                # 1. company_idで探索（最優先）
                normalized_company_id = normalize_company_name(company_id)
                row_idx = company_rows.get(normalized_company_id)
                if row_idx:
                    matched_by = 'company_id'
                
                # 2. company_nameで探索（company_idで見つからない場合）
                if not row_idx and company_id in company_info:
                    company_name = company_info[company_id]['name']
                    normalized_company_name = company_info[company_id]['normalized_name']
                    row_idx = company_rows.get(normalized_company_name)
                    if row_idx:
                        matched_by = 'company_name'
                
                if not row_idx:
                    logger.warning(f"企業が見つかりません（スキップ）: {company_id}")
                    continue
                
                # 各標準品目の価格を転記
                company_filled_count = 0
                company_skipped_count = 0
                written_cells: List[Tuple[str, int, int, float]] = []  # (std_key, row, col, value)のリスト
                
                for std_key, price in items.items():
                    # STDキーを正規化
                    normalized_std_key = normalize_key(std_key)
                    
                    # 正規化後のキーでヘッダーから列番号を取得
                    col_idx = header_to_col.get(normalized_std_key)
                    
                    if not col_idx:
                        company_skipped_count += 1
                        logger.debug(f"  材料が見つかりません（スキップ）: {std_key} (正規化後: {normalized_std_key})")
                        continue
                    
                    # 価格を数値型に変換
                    numeric_price = extract_numeric_price(price)
                    
                    # セルをクリアしてから書き込み
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None  # 一旦クリア
                    
                    if numeric_price is not None:
                        # 数値を整数に変換して書き込み（小数点以下は切り捨て）
                        cell.value = int(numeric_price)
                        company_filled_count += 1
                        total_filled_count += 1
                        written_cells.append((std_key, row_idx, col_idx, numeric_price))
                        
                        # MATCHEDログ出力
                        logger.debug(f"MATCHED: {company_id}, {std_key}, {header_original.get(normalized_std_key, '')}, {numeric_price}")
                
                # 会社ごとのログ出力
                if company_filled_count > 0:
                    logger.info(f"  {company_id}: {company_filled_count}件の価格を転記 (matched_by={matched_by}, スキップ={company_skipped_count})")
                    
                    # 最初の会社のみ、実際に書いたセルのサンプルを10件出力
                    if total_companies == 1 and written_cells:
                        logger.info(f"    サンプル（最初の10件）:")
                        for std_key, row, col, value in written_cells[:10]:
                            logger.info(f"      ({company_id}, {std_key}, row={row}, col={col}, value={value})")
                else:
                    logger.warning(f"  {company_id}: 転記できたセルが0件 (matched_by={matched_by}, スキップ={company_skipped_count})")
            
            # ファイルを保存（数式を保持するため、通常の保存を使用）
            # 既存のシート（特にデバッグ情報シート）の数式を壊さないように注意
            try:
                wb.save(excel_file)
                logger.info(f"✓ {excel_file} - {sheet_name} に転記完了（合計{total_filled_count}件、{total_companies}社）")
                success_count += 1
            except Exception as save_error:
                logger.error(f"Excelファイルの保存エラー: {str(save_error)}")
                # バックアップファイル名で保存を試みる
                backup_file = str(excel_path.with_suffix('.backup.xlsx'))
                try:
                    wb.save(backup_file)
                    logger.warning(f"バックアップファイルに保存しました: {backup_file}")
                except Exception as backup_error:
                    logger.error(f"バックアップ保存も失敗: {str(backup_error)}")
                raise
            
        except Exception as e:
            logger.error(f"エラー: {excel_file} - {sheet_name} - {str(e)}", exc_info=True)
            continue
    
    logger.info(f"\nExcel転記完了: {success_count}/{len([t for t in output_tables if t.get('enabled', True)])} シート")
    return success_count > 0
