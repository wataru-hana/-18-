#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
実装済み企業（現在21社）の価格を自動取得し、Excelに新規シートとして出力するスクリプト
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scrapers import Category1Scraper, Category2Scraper
from scrape_and_fill_standard_table import (
    IMPLEMENTED_COMPANIES,
    MATERIAL_MAPPING,
    normalize_price,
    normalize_company_name,
    load_site_config,
    filter_implemented_companies,
    load_target_items_config,
    load_price_corrections,
    apply_price_corrections,
)

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scrape_log_v2.txt', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# テストシートの材料名の順序（列の順序）
MATERIAL_COLUMNS = [
    'ピカ銅',
    '並銅',
    '砲金',
    '真鍮',
    '雑線80%',
    '雑線60%-65%',
    'VA線',
    'アルミホイール',
    'アルミサッシ',
    'アルミ缶',
    'ステンレス304',
    '鉛バッテリー',
]


def scrape_18_companies():
    """実装済み企業の価格データをスクレイピングで取得"""
    site_configs = load_site_config()
    target_items_config = load_target_items_config()
    price_corrections = load_price_corrections()
    
    # 実装済み企業のみをフィルタリング
    sites = filter_implemented_companies(site_configs)
    
    logger.info(f"実装済み企業: {len(sites)}社を対象にスクレイピングを開始します")
    for i, site in enumerate(sites, 1):
        logger.info(f"  {i}. {site.get('name', '不明')}")
    
    company_results = []
    
    for i, site_config in enumerate(sites, 1):
        company_name = site_config.get('name', '不明')
        category = site_config.get('category', 2)
        normalized_name = site_config.get('normalized_name', normalize_company_name(company_name))
        
        logger.info(f"\n[{i}/{len(sites)}] 処理中: {company_name} (正規化後: {normalized_name})")
        
        try:
            # カテゴリに応じてスクレイパーを選択
            if category == 1:
                scraper = Category1Scraper(site_config, delay=2.0)
            elif category == 2:
                scraper = Category2Scraper(site_config, delay=2.0)
            else:
                logger.warning(f"  不明なカテゴリ: {category}")
                continue
            
            # スクレイピング実行
            result = scraper.scrape(
                filter_target_items=True,
                target_items_config=target_items_config
            )
            
            # 企業名を正規化
            result['company_name'] = normalized_name
            
            company_results.append(result)
            
            # 進捗表示
            prices = result.get('prices', {})
            if prices:
                price_count = len(prices)
                logger.info(f"  ✓ {price_count} 件の価格情報を取得")
            else:
                error = result.get('error', '')
                logger.warning(f"  ✗ 価格情報を取得できませんでした: {error}")
        
        except Exception as e:
            logger.error(f"  エラー: {company_name} - {str(e)}")
            company_results.append({
                'scraped_at': datetime.now().isoformat(),
                'url': site_config.get('price_url', ''),
                'company_name': normalized_name,
                'region': site_config.get('region', ''),
                'error': str(e),
                'prices': {}
            })
    
    # 価格修正マッピングを適用
    if price_corrections:
        logger.info("価格修正マッピングを適用中...")
        company_results = apply_price_corrections(company_results, price_corrections)
    
    return company_results


def save_to_excel_new_sheet(results: List[Dict], excel_file: str = 'price_results_v2_20251104_220253.xlsx', price_corrections: Dict = None):
    """スクレイピング結果をExcelに新規シートとして出力（テストシートと同じ形式）"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # 削除対象の材料を取得
    removed_materials = {}
    if price_corrections:
        for company_name, correction in price_corrections.items():
            if 'remove' in correction:
                removed_materials[company_name] = correction['remove']
    
    # 既存のExcelファイルが存在する場合は、そのファイルに別シートとして追加
    if Path(excel_file).exists():
        wb = load_workbook(excel_file)
        # 新しいシート名を生成（タイムスタンプ付き）
        base_sheet_name = f"テスト_{timestamp}"
        sheet_name = base_sheet_name
        
        # シート名の重複チェック（同じ名前のシートが存在する場合は番号を追加）
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_sheet_name}_{counter}"
            counter += 1
        
        ws = wb.create_sheet(title=sheet_name)
        logger.info(f"既存のExcelファイル '{excel_file}' に新しいシート '{sheet_name}' を追加します")
    else:
        # ファイルが存在しない場合は新規作成
        wb = Workbook()
        ws = wb.active
        ws.title = f"テスト_{timestamp}"
        sheet_name = ws.title
        logger.info(f"新しいExcelファイル '{excel_file}' を作成します")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1行目：ヘッダー行（1列目は空、2列目以降に材料名）
    ws.cell(row=1, column=1, value='').border = border
    for col_idx, material_name in enumerate(MATERIAL_COLUMNS, 2):
        cell = ws.cell(row=1, column=col_idx, value=material_name)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2行目以降：各企業の行（1列目に企業名、2列目以降に各材料の価格）
    row_idx = 2
    for result in results:
        company_name = result.get('company_name', '')
        prices = result.get('prices', {})
        
        # 企業名を1列目に書き込み
        ws.cell(row=row_idx, column=1, value=company_name).border = border
        
        # 各材料の価格をマッピングして書き込み
        normalized_prices = {}
        
        # 削除対象の材料を確認
        company_removed_materials = removed_materials.get(company_name, [])
        
        for material_name, price_value in prices.items():
            # 材料名を正規化
            normalized_material = None
            for key, value in MATERIAL_MAPPING.items():
                if key in material_name or material_name in key:
                    normalized_material = value
                    break
            
            if normalized_material and normalized_material in MATERIAL_COLUMNS:
                # 削除対象の材料かどうかを確認
                should_remove = False
                for removed_material in company_removed_materials:
                    # 材料名を正規化
                    normalized_removed = None
                    for key, value in MATERIAL_MAPPING.items():
                        if key in removed_material or removed_material in key:
                            normalized_removed = value
                            break
                    if not normalized_removed:
                        normalized_removed = removed_material
                    
                    # 正規化後の名前で比較
                    if normalized_material == normalized_removed or removed_material in material_name or material_name in removed_material:
                        should_remove = True
                        break
                
                if should_remove:
                    continue  # 削除対象の材料はスキップ
                
                # 価格を正規化（数値のみ抽出）
                normalized_price = normalize_price(price_value)
                if normalized_price:
                    # 同じ材料で複数の価格がある場合は最初のものを使用
                    if normalized_material not in normalized_prices:
                        normalized_prices[normalized_material] = normalized_price
        
        # 各材料列に価格を書き込み
        for col_idx, material_name in enumerate(MATERIAL_COLUMNS, 2):
            price = normalized_prices.get(material_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=price)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx += 1
    
    # 列幅の自動調整
    ws.column_dimensions['A'].width = 30  # 会社名
    for col_idx in range(2, len(MATERIAL_COLUMNS) + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 15  # 各材料の価格列
    
    # ヘッダー行の高さを設定
    ws.row_dimensions[1].height = 25
    
    wb.save(excel_file)
    logger.info(f"✓ {len(results)}社の取得結果を {excel_file} のシート '{sheet_name}' に保存しました（テストシート形式）")


def main():
    """メイン処理"""
    logger.info("="*80)
    logger.info(f"実装済み{len(IMPLEMENTED_COMPANIES)}社の価格を自動取得してExcelに新規シートとして出力します")
    logger.info("="*80)
    
    # 価格修正マッピングを読み込む
    price_corrections = load_price_corrections()
    
    # スクレイピングを実行
    company_results = scrape_18_companies()
    
    # サマリー表示
    success_count = sum(1 for r in company_results if r.get('prices'))
    total_prices = sum(len(r.get('prices', {})) for r in company_results)
    
    logger.info(f"\n{'='*60}")
    logger.info(f"スクレイピング完了:")
    logger.info(f"  成功: {success_count}/{len(company_results)} 社")
    logger.info(f"  失敗: {len(company_results) - success_count}/{len(company_results)} 社")
    logger.info(f"  取得価格情報総数: {total_prices} 件")
    logger.info(f"{'='*60}")
    
    # Excelに新規シートとして出力
    save_to_excel_new_sheet(company_results, excel_file='プライステスト.xlsx', price_corrections=price_corrections)
    
    logger.info("\n完了しました！")


if __name__ == '__main__':
    main()
