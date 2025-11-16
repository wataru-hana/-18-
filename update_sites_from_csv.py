#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSVファイルから企業情報を読み込んでsites.yamlを更新するスクリプト
"""

import csv
import yaml
import re
from pathlib import Path
from typing import List, Dict
from datetime import datetime

# Excel出力ライブラリのインポート
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxlがインストールされていません。Excel出力機能は使用できません。")

def clean_url(url: str) -> str:
    """URLをクリーンアップ"""
    if not url or url.strip() == '':
        return ''
    url = url.strip()
    # 空の値や無効な値を除外
    if url in ['', ',', ',,']:
        return ''
    return url

def read_csv_companies(csv_path: str) -> List[Dict]:
    """CSVファイルから企業情報を読み込む"""
    companies = []
    
    # エンコーディングを試行
    encodings = ['utf-8', 'shift_jis', 'cp932', 'utf-8-sig']
    
    for encoding in encodings:
        try:
            with open(csv_path, 'r', encoding=encoding, errors='replace') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # 列名の候補を試行（文字化け対応）
                    name = ''
                    region = ''
                    url = ''
                    
                    # 名称を取得（複数の候補を試行）
                    for key in row.keys():
                        if '名称' in key or '名' in key:
                            name = row.get(key, '').strip()
                            break
                    
                    # 地域を取得
                    for key in row.keys():
                        if '地域' in key or '地' in key:
                            region = row.get(key, '').strip()
                            break
                    
                    # URLを取得
                    for key in row.keys():
                        if key == 'URL' or 'URL' in key.upper():
                            url_val = clean_url(row.get(key, ''))
                            if url_val and not url_val.startswith('価格'):
                                url = url_val
                                break
                    
                    # 価格ページURLを収集（列名のパターンを試行）
                    price_urls = []
                    all_keys = list(row.keys())
                    
                    # 価格ページURL列を特定（文字化けを考慮）
                    # 列名のパターン: '価格ペ�ジURL�', '価格ペ�ジURL2', ... など
                    url_patterns = []
                    for key in all_keys:
                        # '価格' または 'URL' を含む列を探す
                        if '価格' in key or ('URL' in key and '価格' in key):
                            url_patterns.append(key)
                    
                    # 番号順にソート（URL1, URL2, ... の順）
                    def get_url_number(key):
                        # 数字を含む場合はその数字を返す
                        import re
                        numbers = re.findall(r'\d+', key)
                        if numbers:
                            return int(numbers[0])
                        # 数字がない場合は1番目と仮定（通常は最初の価格ページURL列）
                        return 1
                    
                    url_patterns.sort(key=get_url_number)
                    
                    # URLを取得
                    for key in url_patterns:
                        price_url = clean_url(row.get(key, ''))
                        if price_url and price_url.startswith('http'):
                            # 重複チェック
                            if price_url not in price_urls:
                                price_urls.append(price_url)
                    
                    if name:  # 名前がある場合のみ追加
                        companies.append({
                            'name': name,
                            'region': region,
                            'url': url,
                            'price_urls': price_urls
                        })
            
            print(f"✓ CSVファイルを {encoding} エンコーディングで読み込みました: {len(companies)} 社")
            return companies
            
        except Exception as e:
            print(f"  {encoding} で読み込み失敗: {e}")
            continue
    
    return companies

def load_existing_sites(yaml_path: str) -> List[Dict]:
    """既存のsites.yamlを読み込む"""
    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('sites', [])
    except FileNotFoundError:
        return []
    except Exception as e:
        print(f"既存のsites.yamlの読み込みエラー: {e}")
        return []

def normalize_company_name(name: str) -> str:
    """企業名を正規化（比較用）"""
    # 括弧内の文字を削除して比較
    name = re.sub(r'[（(].*?[）)]', '', name)
    name = name.strip()
    # 全角スペースを半角に
    name = name.replace('　', ' ')
    # 連続するスペースを1つに
    name = re.sub(r'\s+', ' ', name)
    return name

def is_company_exists(company_name: str, existing_sites: List[Dict]) -> bool:
    """企業が既に登録されているか確認"""
    normalized_new = normalize_company_name(company_name)
    
    for site in existing_sites:
        existing_name = site.get('name', '')
        normalized_existing = normalize_company_name(existing_name)
        
        # 完全一致または部分一致を確認
        if (normalized_new == normalized_existing or 
            normalized_new in normalized_existing or 
            normalized_existing in normalized_new):
            return True
    
    return False

def create_site_config(company: Dict) -> Dict:
    """企業情報からサイト設定を作成"""
    config = {
        'name': company['name'],
        'category': 2,  # デフォルトはカテゴリ2
        'region': company['region'],
        'extractor_type': 'auto'  # デフォルトは自動抽出
    }
    
    # 価格URLの設定
    if len(company['price_urls']) == 1:
        config['price_url'] = company['price_urls'][0]
    elif len(company['price_urls']) > 1:
        config['price_url'] = company['price_urls'][0]  # メインURL
        config['price_urls'] = company['price_urls']  # 複数URL
    
    return config

def update_excel_companies_sheet(excel_path: str, companies: List[Dict], new_companies: List[Dict]):
    """
    Excelファイルに企業情報シートを追加または更新
    
    Args:
        excel_path: Excelファイルのパス
        companies: 全企業情報のリスト
        new_companies: 新規追加された企業のリスト
    """
    if not EXCEL_AVAILABLE:
        print("  ⚠ Excel出力機能は使用できません（openpyxlが必要）")
        return
    
    try:
        # 既存のExcelファイルを読み込むか、新規作成
        if Path(excel_path).exists():
            wb = load_workbook(excel_path)
            # 既存の「企業一覧」シートがあれば削除
            if '企業一覧' in wb.sheetnames:
                wb.remove(wb['企業一覧'])
            ws = wb.create_sheet(title='企業一覧', index=0)  # 最初のシートとして追加
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = '企業一覧'
        
        # ヘッダーの設定
        headers = ['会社名', '地域', 'URL', '価格ページURL1', '価格ページURL2', '価格ページURL3', 
                   '価格ページURL4', '価格ページURL5', '価格ページURL6', '追加日時']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行を書き込み
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # データ行を書き込み
        row_idx = 2
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        for company in companies:
            # 新規追加企業かどうかを確認
            is_new = any(nc['name'] == company['name'] for nc in new_companies)
            added_time = timestamp if is_new else ''
            
            ws.cell(row=row_idx, column=1, value=company['name']).border = border
            ws.cell(row=row_idx, column=2, value=company.get('region', '')).border = border
            ws.cell(row=row_idx, column=3, value=company.get('url', '')).border = border
            
            # 価格ページURLを書き込み
            price_urls = company.get('price_urls', [])
            if not price_urls and company.get('price_url'):
                price_urls = [company['price_url']]
            
            for i, url in enumerate(price_urls[:6], 1):  # 最大6つまで
                ws.cell(row=row_idx, column=3 + i, value=url).border = border
            
            ws.cell(row=row_idx, column=10, value=added_time).border = border
            
            # 新規追加企業の行をハイライト
            if is_new:
                for col in range(1, 11):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            
            row_idx += 1
        
        # 列幅の自動調整
        column_widths = {
            'A': 35,  # 会社名
            'B': 12,  # 地域
            'C': 50,  # URL
            'D': 50,  # 価格ページURL1
            'E': 50,  # 価格ページURL2
            'F': 50,  # 価格ページURL3
            'G': 50,  # 価格ページURL4
            'H': 50,  # 価格ページURL5
            'I': 50,  # 価格ページURL6
            'J': 20   # 追加日時
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # ヘッダー行の高さを設定
        ws.row_dimensions[1].height = 25
        
        # データ行の文字列折り返し設定
        for row in ws.iter_rows(min_row=2, max_row=row_idx-1):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        wb.save(excel_path)
        print(f"  ✓ Excelファイルに企業情報シートを追加しました: {excel_path}")
        
    except Exception as e:
        print(f"  ⚠ Excelファイルの更新に失敗しました: {str(e)}")

def main():
    """メイン処理"""
    csv_path = '非鉄金属業者一覧最終提出用 - CSV.csv'
    yaml_path = 'config/sites.yaml'
    
    print("=" * 60)
    print("CSVファイルから企業情報を読み込んでsites.yamlを更新します")
    print("=" * 60)
    
    # CSVファイルから企業情報を読み込む
    csv_companies = read_csv_companies(csv_path)
    if not csv_companies:
        print("エラー: CSVファイルから企業情報を読み込めませんでした")
        return
    
    # 既存のsites.yamlを読み込む
    existing_sites = load_existing_sites(yaml_path)
    existing_names = [site.get('name', '') for site in existing_sites]
    
    print(f"\n既存の登録企業数: {len(existing_sites)}")
    print(f"CSVファイルの企業数: {len(csv_companies)}")
    
    # 新しい企業を特定し、既存企業のURLを更新
    new_companies = []
    updated_companies = []
    skipped_companies = []
    
    for company in csv_companies:
        # 既存企業かどうかを確認
        existing_site = None
        for site in existing_sites:
            if normalize_company_name(site.get('name', '')) == normalize_company_name(company['name']):
                existing_site = site
                break
        
        if existing_site:
            # 既存企業のURLを更新（URLが空または不足している場合）
            existing_price_urls = existing_site.get('price_urls', [])
            existing_price_url = existing_site.get('price_url', '')
            
            # 既存のURLを収集
            if existing_price_url:
                if existing_price_url not in existing_price_urls:
                    existing_price_urls.insert(0, existing_price_url)
            
            # CSVから取得したURLと比較
            csv_price_urls = company.get('price_urls', [])
            
            # URLが更新されている場合
            if csv_price_urls and (not existing_price_urls or set(csv_price_urls) != set(existing_price_urls)):
                updated_companies.append({
                    'name': company['name'],
                    'existing': existing_site,
                    'new_urls': csv_price_urls
                })
            else:
                skipped_companies.append(company['name'])
        else:
            new_companies.append(company)
    
    print(f"\n新規追加対象: {len(new_companies)} 社")
    print(f"URL更新対象: {len(updated_companies)} 社")
    print(f"変更なし: {len(skipped_companies)} 社")
    
    if skipped_companies:
        print("\n変更なしの企業:")
        for name in skipped_companies[:10]:  # 最初の10社のみ表示
            print(f"  - {name}")
        if len(skipped_companies) > 10:
            print(f"  ... 他 {len(skipped_companies) - 10} 社")
    
    if not new_companies and not updated_companies:
        print("\n追加・更新する企業はありません。")
        # Excelファイルへの追加のみ実行（既存企業情報をExcelに記録）
        excel_path = 'price_results_v2_20251104_220253.xlsx'
        if Path(excel_path).exists():
            print(f"\nExcelファイルに企業情報シートを追加中...")
            # 全企業情報を準備（sites.yamlから）
            all_companies_data = []
            for site in existing_sites:
                company_data = {
                    'name': site.get('name', ''),
                    'region': site.get('region', ''),
                    'url': site.get('url', ''),
                    'price_url': site.get('price_url', ''),
                    'price_urls': site.get('price_urls', [])
                }
                all_companies_data.append(company_data)
            
            update_excel_companies_sheet(excel_path, all_companies_data, [])
        return
    
    # 新しい企業を設定に追加
    if new_companies:
        print(f"\n新規追加企業:")
        for company in new_companies:
            print(f"  - {company['name']} ({company['region']})")
            if company['price_urls']:
                print(f"    URL数: {len(company['price_urls'])}")
                for i, url in enumerate(company['price_urls'][:3], 1):
                    print(f"      {i}. {url[:60]}...")
                if len(company['price_urls']) > 3:
                    print(f"      ... 他 {len(company['price_urls']) - 3} URL")
            else:
                print(f"    ⚠ 価格ページURLが設定されていません")
    
    # URL更新対象企業
    if updated_companies:
        print(f"\nURL更新対象企業:")
        for item in updated_companies:
            print(f"  - {item['name']}")
            print(f"    既存URL数: {len(item['existing'].get('price_urls', []) + ([item['existing'].get('price_url', '')] if item['existing'].get('price_url') else []))}")
            print(f"    新規URL数: {len(item['new_urls'])}")
            for i, url in enumerate(item['new_urls'][:3], 1):
                print(f"      {i}. {url[:60]}...")
            if len(item['new_urls']) > 3:
                print(f"      ... 他 {len(item['new_urls']) - 3} URL")
    
    # バックアップを作成
    backup_path = f"{yaml_path}.backup"
    if Path(yaml_path).exists():
        import shutil
        shutil.copy2(yaml_path, backup_path)
        print(f"\n既存の設定ファイルをバックアップしました: {backup_path}")
    
    # 設定を更新
    updated_sites = []
    
    # 既存のサイトを更新
    for site in existing_sites:
        # URL更新対象かどうかを確認
        updated_item = None
        for item in updated_companies:
            if normalize_company_name(site.get('name', '')) == normalize_company_name(item['name']):
                updated_item = item
                break
        
        if updated_item:
            # URLを更新
            site = site.copy()
            csv_company = None
            for c in csv_companies:
                if normalize_company_name(c['name']) == normalize_company_name(updated_item['name']):
                    csv_company = c
                    break
            
            if csv_company and csv_company.get('price_urls'):
                if len(csv_company['price_urls']) == 1:
                    site['price_url'] = csv_company['price_urls'][0]
                    if 'price_urls' in site:
                        del site['price_urls']
                elif len(csv_company['price_urls']) > 1:
                    site['price_url'] = csv_company['price_urls'][0]
                    site['price_urls'] = csv_company['price_urls']
            
            # 地域も更新（CSVに含まれている場合）
            if csv_company and csv_company.get('region'):
                site['region'] = csv_company['region']
        
        updated_sites.append(site)
    
    # 新しい企業を追加
    for company in new_companies:
        site_config = create_site_config(company)
        updated_sites.append(site_config)
    
    # YAMLファイルに書き込み
    config = {'sites': updated_sites}
    with open(yaml_path, 'w', encoding='utf-8') as f:
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    
    print(f"\n✓ sites.yamlを更新しました")
    print(f"  総企業数: {len(updated_sites)} 社")
    print(f"  新規追加: {len(new_companies)} 社")
    print(f"  URL更新: {len(updated_companies)} 社")
    
    # Excelファイルに企業情報を追加
    excel_path = 'price_results_v2_20251104_220253.xlsx'
    if Path(excel_path).exists() or new_companies or updated_companies:
        print(f"\nExcelファイルを更新中...")
        # 全企業情報を準備（sites.yamlから）
        all_companies_data = []
        for site in updated_sites:
            company_data = {
                'name': site.get('name', ''),
                'region': site.get('region', ''),
                'url': site.get('url', ''),
                'price_url': site.get('price_url', ''),
                'price_urls': site.get('price_urls', [])
            }
            all_companies_data.append(company_data)
        
        update_excel_companies_sheet(excel_path, all_companies_data, new_companies)
    
    print("\n次のステップ:")
    print("1. config/sites.yamlを確認して、必要に応じて設定を調整してください")
    print("2. 各企業のHTML構造を分析して、適切なextractor_typeを設定してください")
    print("3. scrape_prices_v2.pyを実行してスクレイピングをテストしてください")

if __name__ == '__main__':
    main()

