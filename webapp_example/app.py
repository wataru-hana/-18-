#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
価格自動取得システム - Flask Webアプリケーション例
"""

from flask import Flask, render_template, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import sys
import os
import re

# 既存のスクレイパーモジュールをインポート
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, parent_dir)
print(f"DEBUG: Parent directory: {parent_dir}")
print(f"DEBUG: Scrapers path exists: {os.path.exists(os.path.join(parent_dir, 'scrapers'))}")
print(f"DEBUG: Scrapers __init__.py exists: {os.path.exists(os.path.join(parent_dir, 'scrapers', '__init__.py'))}")

try:
    from scrapers import Category1Scraper, Category2Scraper
    print("DEBUG: Successfully imported scrapers")
except ImportError as e:
    print(f"DEBUG: Import error: {e}")
    import traceback
    traceback.print_exc()
    raise
import yaml

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///prices.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# グローバルエラーハンドラー（すべてのエラーをJSONで返す）
@app.errorhandler(Exception)
def handle_exception(e):
    """すべての例外をJSONで返す"""
    import traceback
    return jsonify({
        'status': 'error',
        'error': str(e),
        'type': type(e).__name__,
        'trace': traceback.format_exc()
    }), 500

@app.errorhandler(500)
def handle_500(e):
    """500エラーをJSONで返す"""
    import traceback
    return jsonify({
        'status': 'error',
        'error': 'Internal Server Error',
        'details': str(e),
        'trace': traceback.format_exc()
    }), 500

# データベースモデル
class Company(db.Model):
    """企業モデル"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    region = db.Column(db.String(50))
    price_url = db.Column(db.String(500))
    category = db.Column(db.Integer)
    extractor_type = db.Column(db.String(50))
    is_implemented = db.Column(db.Boolean, default=False)
    
    prices = db.relationship('PriceData', backref='company', lazy=True)

class PriceData(db.Model):
    """価格データモデル"""
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=False)
    material_name = db.Column(db.String(100), nullable=False)
    price = db.Column(db.String(50))
    scraped_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'company': self.company.name,
            'material': self.material_name,
            'price': self.price,
            'scraped_at': self.scraped_at.isoformat() if self.scraped_at else None
        }

# ルーティング
@app.route('/')
def index():
    """メインページ"""
    companies = Company.query.filter_by(is_implemented=True).all()
    return render_template('index.html', companies=companies)

@app.route('/api/companies')
def get_companies():
    """企業一覧を取得"""
    companies = Company.query.filter_by(is_implemented=True).all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'region': c.region
    } for c in companies])

@app.route('/api/scrape', methods=['POST'])
def start_scraping():
    """スクレイピングを開始（同期的に実行）"""
    data = request.json or {}
    company_ids = data.get('company_ids', None)
    
    # 設定ファイルのパスを取得（webapp_example内または親ディレクトリから）
    def get_config_path(filename):
        """設定ファイルのパスを取得（デプロイ環境に対応）"""
        # まずwebapp_example内のconfigフォルダを確認
        local_path = os.path.join(os.path.dirname(__file__), 'config', filename)
        if os.path.exists(local_path):
            return local_path
        # 次に親ディレクトリのconfigフォルダを確認
        parent_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', filename)
        if os.path.exists(parent_path):
            return parent_path
        # どちらも見つからない場合は親ディレクトリを返す
        return parent_path
    
    # 設定ファイルを読み込み
    config_path = get_config_path('sites.yaml')
    with open(config_path, 'r', encoding='utf-8') as f:
        sites_config = yaml.safe_load(f)
        sites = sites_config.get('sites', [])
    
    # 対象アイテム設定を読み込み
    target_items_path = get_config_path('target_items.yaml')
    with open(target_items_path, 'r', encoding='utf-8') as f:
        target_items_config = yaml.safe_load(f)
        target_items = target_items_config.get('target_items', [])
    
    # 価格修正マッピングを読み込み
    corrections_path = get_config_path('price_corrections.yaml')
    with open(corrections_path, 'r', encoding='utf-8') as f:
        corrections_config = yaml.safe_load(f)
        corrections = corrections_config.get('corrections', {})
    
    results = []
    
    # 実装済み21社のリスト（3社追加: 鴻陽産業、大垣金属、高橋商事）
    IMPLEMENTED_COMPANIES = {
        '眞田鋼業株式会社', '有限会社金田商事', '木村金属（大阪）',
        '明鑫貿易株式会社', '東起産業（株）', '土金（大阪）',
        '大畑商事（千葉・大阪）', '千福商会（大阪）', '鴻祥貿易株式会社',
        '株式会社鳳山', '株式会社 春日商会　富山支店',
        '株式会社 春日商会　滋賀支店', '株式会社 春日商会　一宮本社',
        '安城貿易（愛知）', '東北キング', '株式会社八木',
        '有限会社　八尾アルミセンター', '株式会社 ヒラノヤ',
        '鴻陽産業株式会社 岐阜工場', '株式会社 大垣金属',
        '高橋商事株式会社'
    }
    
    # 実装済み企業のみをフィルタリング
    implemented_sites = []
    for site in sites:
        company_name = site.get('name', '')
        # 文字化けなどを補正して正規化した名前を取得
        company_name_normalized = normalize_company_name(company_name)
        
        # 正規化後の企業名が実装済み18社に含まれているか確認
        # 完全一致または部分一致で確認
        is_implemented = False
        if company_name_normalized in IMPLEMENTED_COMPANIES:
            is_implemented = True
        else:
            # 部分一致で確認（文字化けなどで完全一致しない場合のフォールバック）
            for impl_name in IMPLEMENTED_COMPANIES:
                # 正規化した実装済み企業名と比較
                impl_normalized = normalize_company_name(impl_name)
                if (impl_normalized in company_name_normalized or 
                    company_name_normalized in impl_normalized or
                    impl_name in company_name_normalized or
                    company_name_normalized in impl_name):
                    is_implemented = True
                    break
        
        if is_implemented:
            implemented_sites.append(site)
    
    # スクレイピング実行
    for site_config in implemented_sites:
        company_name = site_config.get('name', '不明')
        category = site_config.get('category', 2)
        
        try:
            # カテゴリに応じてスクレイパーを選択
            if category == 1:
                scraper = Category1Scraper(site_config, delay=2.0)
            elif category == 2:
                scraper = Category2Scraper(site_config, delay=2.0)
            else:
                continue
            
            # スクレイピング実行
            result = scraper.scrape(
                filter_target_items=True,
                target_items_config=target_items
            )
            
            # デバッグ: resultの内容を確認
            print(f"DEBUG {company_name}: result type={type(result)}, keys={result.keys() if isinstance(result, dict) else 'N/A'}")
            if isinstance(result, dict):
                print(f"DEBUG {company_name}: prices type={type(result.get('prices'))}, value={result.get('prices')}")
            
            # 企業名を正規化
            company_name_normalized = normalize_company_name(company_name)
            result['company_name'] = company_name_normalized
            
            # 価格修正マッピングを適用
            if company_name_normalized in corrections:
                try:
                    correction_config = corrections[company_name_normalized]
                    if not isinstance(correction_config, dict):
                        print(f"WARNING: correction_config is not a dict for {company_name_normalized}: {type(correction_config)}")
                    else:
                        result = apply_price_corrections_single(result, correction_config)
                except Exception as e:
                    print(f"ERROR in apply_price_corrections_single for {company_name_normalized}: {e}")
                    import traceback
                    traceback.print_exc()
                    # エラーが発生しても続行（修正なしで続ける）
            
            # データベースに保存
            company = Company.query.filter_by(name=company_name_normalized).first()
            if not company:
                company = Company(
                    name=company_name_normalized,
                    region=site_config.get('region', ''),
                    price_url=site_config.get('price_url', ''),
                    category=category,
                    extractor_type=site_config.get('extractor_type', ''),
                    is_implemented=True
                )
                db.session.add(company)
                db.session.commit()
            
            # 価格データを保存
            prices = result.get('prices', {})
            if not isinstance(prices, dict):
                raise ValueError(f"prices is not a dict: {type(prices)}, value={prices}")
            for material_name, price_value in prices.items():
                price_data = PriceData(
                    company_id=company.id,
                    material_name=material_name,
                    price=price_value,
                    scraped_at=datetime.utcnow()
                )
                db.session.add(price_data)
            
            results.append({
                'company': company_name_normalized,
                'price_count': len(prices),
                'status': 'success'
            })
        
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"エラー発生: {company_name}")
            print(f"エラー内容: {str(e)}")
            print(f"詳細: {error_detail}")
            results.append({
                'company': company_name,
                'status': 'error',
                'error': f"{str(e)}: {error_detail[:200]}"
            })
    
    db.session.commit()
    
    # 各材料の最高価格を計算
    max_prices = calculate_max_prices()
    
    return jsonify({
        'status': 'completed',
        'results': results,
        'total': len(results),
        'max_prices': max_prices
    })

@app.route('/api/reset-database', methods=['POST'])
def reset_database():
    """データベースをリセット（全データ削除）"""
    try:
        # 全ての価格データを削除
        PriceData.query.delete()
        # 全ての会社データを削除
        Company.query.delete()
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Database reset complete'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/results')
def get_results():
    """最新の結果を取得"""
    latest_scrape_time = db.session.query(db.func.max(PriceData.scraped_at)).scalar()
    
    if not latest_scrape_time:
        return jsonify([])
    
    results = PriceData.query.filter_by(scraped_at=latest_scrape_time).all()
    return jsonify([r.to_dict() for r in results])

@app.route('/api/results/latest')
def get_latest_results():
    """企業ごとの最新価格を取得"""
    companies = Company.query.filter_by(is_implemented=True).all()
    results = []
    
    for company in companies:
        latest_price = PriceData.query.filter_by(company_id=company.id)\
            .order_by(PriceData.scraped_at.desc()).first()
        
        if latest_price:
            # その企業の最新の全価格を取得
            latest_scrape_time = db.session.query(db.func.max(PriceData.scraped_at))\
                .filter_by(company_id=company.id).scalar()
            
            prices = PriceData.query.filter_by(
                company_id=company.id,
                scraped_at=latest_scrape_time
            ).all()
            
            results.append({
                'company': company.name,
                'region': company.region,
                'prices': {p.material_name: p.price for p in prices},
                'scraped_at': latest_scrape_time.isoformat() if latest_scrape_time else None
            })
    
    return jsonify(results)

@app.route('/api/results/max-prices')
def get_max_prices():
    """各材料の最高価格を取得"""
    max_prices = calculate_max_prices()
    return jsonify(max_prices)

def calculate_max_prices():
    """各材料の最高価格を計算"""
    # 最新のスクレイピング時刻を取得
    latest_scrape_time = db.session.query(db.func.max(PriceData.scraped_at)).scalar()
    
    if not latest_scrape_time:
        return []
    
    # 最新のスクレイピング結果のみを取得
    latest_prices = PriceData.query.filter_by(scraped_at=latest_scrape_time).all()
    
    # 材料ごとに価格を集計
    material_prices = {}
    for price_data in latest_prices:
        material_name = price_data.material_name
        price_str = price_data.price
        
        # 価格文字列から数値を抽出
        price_value = extract_price_number(price_str)
        
        if price_value is not None:
            if material_name not in material_prices:
                material_prices[material_name] = []
            
            material_prices[material_name].append({
                'company': price_data.company.name,
                'region': price_data.company.region,
                'price': price_value,
                'price_str': price_str
            })
    
    # 各材料の最高価格を計算
    max_prices_list = []
    for material_name, prices in material_prices.items():
        if prices:
            # 最高価格を取得
            max_price_item = max(prices, key=lambda x: x['price'])
            max_prices_list.append({
                'material': material_name,
                'max_price': max_price_item['price_str'],
                'max_price_value': max_price_item['price'],
                'company': max_price_item['company'],
                'region': max_price_item['region']
            })
    
    # 材料名でソート
    max_prices_list.sort(key=lambda x: x['material'])
    
    return max_prices_list

def extract_price_number(price_str):
    """価格文字列から数値を抽出"""
    if not price_str:
        return None
    
    # 数値を抽出（カンマや円マークを除去）
    import re
    price_match = re.search(r'(\d{1,4}(?:[,，]\d{3})*(?:\.\d+)?)', str(price_str))
    if price_match:
        price_value = price_match.group(1).replace(',', '').replace('，', '')
        try:
            return float(price_value)
        except ValueError:
            return None
    return None

@app.route('/api/download/excel')
def download_excel():
    """Excelファイルをダウンロード（添付画像の形式で出力）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    wb = Workbook()
    
    # メインシート：添付画像の形式で価格一覧表を作成
    ws_table = wb.active
    ws_table.title = "価格一覧表"
    
    # 21社のリスト（3社追加: 鴻陽産業、大垣金属、高橋商事）
    COMPANY_LIST = [
        '眞田鋼業株式会社',
        '有限会社金田商事',
        '木村金属（大阪）',
        '明鑫貿易株式会社',
        '東起産業（株）',
        '土金（大阪）',
        '大畑商事（千葉・大阪）',
        '千福商会（大阪）',
        '鴻祥貿易株式会社',
        '株式会社鳳山',
        '株式会社 春日商会　富山支店',
        '株式会社 春日商会　滋賀支店',
        '株式会社 春日商会　一宮本社',
        '安城貿易（愛知）',
        '東北キング',
        '株式会社八木',
        '有限会社　八尾アルミセンター',
        '株式会社 ヒラノヤ',
        '鴻陽産業株式会社 岐阜工場',
        '株式会社 大垣金属',
        '高橋商事株式会社'
    ]
    
    # 品目リスト（添付画像の順番）
    MATERIAL_LIST = [
        'ピカ銅',
        '並銅',
        '砲金',
        '真鍮',
        '雑線80%',
        '雑60%-65%',
        'VA線',
        'アルミホイール',
        'アルミサッシ',
        'アルミ缶',
        'バラアルミ缶',
        'プレスステンレス304',
        '鉛バッテリー'
    ]
    
    # 材料名のマッピング（スクレイピング結果と表ヘッダーを一致させる）
    MATERIAL_MAPPING = {
        # ピカ銅
        'ピカ銅': 'ピカ銅', 'ピカ線': 'ピカ銅', 'ピカドウ': 'ピカ銅',
        '1号銅': 'ピカ銅', '一号銅': 'ピカ銅', '特一号銅': 'ピカ銅',
        '上銅': 'ピカ銅',  # 土金
        
        # 並銅
        '並銅': '並銅', '波銅': '並銅', '波道': '並銅', '2号銅': '並銅',
        '込銅': '並銅',  # 鴻祥貿易
        
        # 砲金
        '砲金': '砲金', 'ほうきん': '砲金', 'gunmetal': '砲金',
        
        # 真鍮
        '真鍮': '真鍮', 'しんちゅう': '真鍮', '黄銅': '真鍮',
        '真鍮A': '真鍮',  # 東起産業
        '真鍮（上）': '真鍮',  # 土金
        '真鍮B': '真鍮',  # 鴻祥貿易
        
        # 雑線80%
        '雑線80%': '雑線80%', '雑電線80%': '雑線80%', '電線80%': '雑線80%', '雑線（80%）': '雑線80%',
        '一本線80%': '雑線80%',  # 東起産業
        '雑線S': '雑線80%',  # 土金
        '上線': '雑線80%',  # 鴻祥貿易
        
        # 雑線60%-65%
        '雑線60%': '雑60%-65%', '雑線65%': '雑60%-65%', '雑線60%-65%': '雑60%-65%',
        '雑電線60%': '雑60%-65%', '電線60%': '雑60%-65%', '雑線（60%）': '雑60%-65%',
        '雑60%-65%': '雑60%-65%',
        '三本線65%': '雑60%-65%',  # 東起産業
        '雑線A': '雑60%-65%',  # 土金
        
        # VA線
        'VA線': 'VA線', 'VVF': 'VA線', 'VVFケーブル': 'VA線', 'ＶＡ線': 'VA線',
        'ＶＡ線(巻き)': 'VA線',  # 土金
        'ねずみ線': 'VA線',  # 鴻祥貿易
        
        # アルミホイール
        'アルミホイール': 'アルミホイール', 'ホイール': 'アルミホイール', 'Alホイール': 'アルミホイール',
        
        # アルミサッシ
        'アルミサッシ': 'アルミサッシ', 'サッシ': 'アルミサッシ', 'Alサッシ': 'アルミサッシ',
        'アルミサッシA 付物なし': 'アルミサッシ',  # 東起産業
        'アルミサッシA': 'アルミサッシ',  # 鴻祥貿易
        
        # アルミ缶
        'アルミ缶': 'アルミ缶', 'アルミ缶バラ': 'アルミ缶', '缶バラ': 'アルミ缶', 'アルミ缶　バラ': 'アルミ缶',
        
        # バラアルミ缶（プレス）
        'アルミ缶プレス': 'バラアルミ缶', '缶プレス': 'バラアルミ缶', 'アルミ缶　プレス': 'バラアルミ缶', 'バラアルミ缶': 'バラアルミ缶',
        'アルミ缶（プレス）': 'バラアルミ缶',  # 土金
        
        # ステンレス304
        'SUS304': 'プレスステンレス304', 'ステンレス304': 'プレスステンレス304', '304': 'プレスステンレス304',
        'ステン304': 'プレスステンレス304', 'SUS': 'プレスステンレス304', 'プレスステンレス304': 'プレスステンレス304',
        'ステンレス': 'プレスステンレス304',  # 東起産業
        'ステンレス 304': 'プレスステンレス304',  # 鴻祥貿易
        'ステンレス（上）': 'プレスステンレス304',  # 土金
        
        # 鉛バッテリー
        '鉛バッテリー': '鉛バッテリー', 'バッテリー': '鉛バッテリー', '鉛': '鉛バッテリー',
        '自動車バッテリー': '鉛バッテリー',  # 東起産業
        'バッテリー（上）': '鉛バッテリー',  # 土金
    }
    
    # スタイル設定
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # ヘッダー行（1行目）：1列目は空（会社名列）、2列目以降が品目名
    ws_table.cell(row=1, column=1, value='').border = thin_border
    for col_idx, material in enumerate(MATERIAL_LIST, 2):
        cell = ws_table.cell(row=1, column=col_idx, value=material)
        cell.alignment = center_align
        cell.border = thin_border
    
    # 会社名列（1列目、2行目以降）と空セルに罫線
    for row_idx, company in enumerate(COMPANY_LIST, 2):
        cell = ws_table.cell(row=row_idx, column=1, value=company)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        # 品目列に罫線を追加（空セル）
        for col_idx in range(2, len(MATERIAL_LIST) + 2):
            ws_table.cell(row=row_idx, column=col_idx).border = thin_border
    
    # データベースから全ての価格データを取得（材料名・企業名のペアごとに最新のものを使用）
    all_prices = PriceData.query.order_by(PriceData.scraped_at.desc()).all()
    
    # 企業名→材料名→価格のディクショナリを作成（最新のもののみ保持）
    price_dict = {}  # {normalized_company_name: {normalized_material: price}}
    
    for price_data in all_prices:
        company = price_data.company
        if not company:
            continue
        
        company_name_normalized = normalize_company_name(company.name)
        material_name = price_data.material_name
        
        # 材料名を正規化
        normalized_material = None
        for key, value in MATERIAL_MAPPING.items():
            if key in material_name or material_name in key:
                normalized_material = value
                break
        
        if not normalized_material:
            continue
        
        # まだこの組み合わせの価格がなければ追加（降順なので最初が最新）
        if company_name_normalized not in price_dict:
            price_dict[company_name_normalized] = {}
        
        if normalized_material not in price_dict[company_name_normalized]:
            price_value = normalize_price(price_data.price)
            if price_value:
                price_dict[company_name_normalized][normalized_material] = price_value
    
    # 表に価格を記入
    for row_idx, table_company in enumerate(COMPANY_LIST, 2):
        table_company_normalized = normalize_company_name(table_company)
        
        # price_dictから該当する企業を探す
        matched_company = None
        for dict_company in price_dict.keys():
            if (table_company_normalized == dict_company or
                table_company_normalized in dict_company or
                dict_company in table_company_normalized):
                matched_company = dict_company
                break
        
        if matched_company is None:
            continue
        
        # 各材料の価格を記入
        for col_idx, table_material in enumerate(MATERIAL_LIST, 2):
            if table_material in price_dict[matched_company]:
                price_value = price_dict[matched_company][table_material]
                try:
                    cell = ws_table.cell(row=row_idx, column=col_idx, value=int(price_value))
                    cell.alignment = center_align
                    cell.border = thin_border
                except (ValueError, TypeError):
                    cell = ws_table.cell(row=row_idx, column=col_idx, value=str(price_value))
                    cell.alignment = center_align
                    cell.border = thin_border
    
    # 列幅を調整
    ws_table.column_dimensions['A'].width = 28  # 会社名列
    for col_idx in range(2, len(MATERIAL_LIST) + 2):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
        ws_table.column_dimensions[col_letter].width = 12
    
    # デバッグ用シート：データベースに保存されている全価格データを表示
    ws_debug = wb.create_sheet("デバッグ情報")
    debug_headers = ['企業名', '材料名', '価格', '取得日時', '正規化企業名', '正規化材料名']
    for col_idx, header in enumerate(debug_headers, 1):
        ws_debug.cell(row=1, column=col_idx, value=header)
    
    debug_row = 2
    for price_data in all_prices[:500]:  # 最大500件
        company = price_data.company
        if company:
            company_name = company.name
            company_normalized = normalize_company_name(company.name)
        else:
            company_name = "不明"
            company_normalized = "不明"
        
        material_name = price_data.material_name
        
        # 材料名を正規化
        normalized_material = "マッピングなし"
        for key, value in MATERIAL_MAPPING.items():
            if key in material_name or material_name in key:
                normalized_material = value
                break
        
        ws_debug.cell(row=debug_row, column=1, value=company_name)
        ws_debug.cell(row=debug_row, column=2, value=material_name)
        ws_debug.cell(row=debug_row, column=3, value=price_data.price)
        ws_debug.cell(row=debug_row, column=4, value=str(price_data.scraped_at) if price_data.scraped_at else '')
        ws_debug.cell(row=debug_row, column=5, value=company_normalized)
        ws_debug.cell(row=debug_row, column=6, value=normalized_material)
        debug_row += 1
    
    # メモリに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'price_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

def normalize_company_name(name):
    """企業名を正規化（scrape_18_companies_to_excel.pyと同じロジック）"""
    if not name:
        return ''
    
    # 正規化後の統一名称（Excel出力用）
    COMPANY_NAME_MAPPING = {
        # 正常な会社名
        '眞田鋼業株式会社': '眞田鋼業株式会社',
        '明鑫貿易株式会社': '明鑫貿易株式会社',
        '東起産業（株）': '東起産業（株）',
        '鴻祥貿易株式会社': '鴻祥貿易株式会社',
        '安城貿易（愛知）': '安城貿易（愛知）',
        '千福商会（大阪）': '千福商会（大阪）',
        '土金（大阪）': '土金（大阪）',
        '大畑商事（千葉・大阪）': '大畑商事（千葉・大阪）',
        '木村金属（大阪）': '木村金属（大阪）',
        '株式会社 春日商会　富山支店': '株式会社 春日商会　富山支店',
        '株式会社 春日商会　滋賀支店': '株式会社 春日商会　滋賀支店',
        '株式会社 春日商会　一宮本社': '株式会社 春日商会　一宮本社',
        '株式会社八木': '株式会社八木',
        '有限会社　八尾アルミセンター': '有限会社　八尾アルミセンター',
        '株式会社 ヒラノヤ': '株式会社 ヒラノヤ',
        '株式会社鳳山': '株式会社鳳山',
        '東北キング': '東北キング',
        '有限会社金田商事': '有限会社金田商事',
        # 追加企業（2024年11月30日）
        '鴻陽産業株式会社 岐阜工場': '鴻陽産業株式会社 岐阜工場',
        '鴻陽産業株式会社': '鴻陽産業株式会社 岐阜工場',
        '双王金属': '鴻陽産業株式会社 岐阜工場',
        '株式会社 大垣金属': '株式会社 大垣金属',
        '大垣金属': '株式会社 大垣金属',
        '高橋商事株式会社': '高橋商事株式会社',
        '高橋商事': '高橋商事株式会社',
        # 文字化けバージョン -> 正常な会社名へのマッピング
        '明鑫貿易�式会社': '明鑫貿易株式会社',
        '東起産業��檼': '東起産業（株）',
        '東起産業檼': '東起産業（株）',
        '鴻祥貿易�式会社': '鴻祥貿易株式会社',
        '安城貿易（�知�': '安城貿易（愛知）',
        '卦�商会（大阪�': '千福商会（大阪）',
        '土�߼�大阪�': '土金（大阪）',
        '土金�大阪�': '土金（大阪）',
        '大畑商事（千葉�大阪�': '大畑商事（千葉・大阪）',
        '木村��属（大阪�': '木村金属（大阪）',
        '木村金属（大阪�': '木村金属（大阪）',
        '株式会社 春日啼 富山支�': '株式会社 春日商会　富山支店',
        '株式会社 春日啼 滋�支�': '株式会社 春日商会　滋賀支店',
        '株式会社 春日啼 �宮本社': '株式会社 春日商会　一宮本社',
    }
    
    name = str(name).strip()
    
    # マッピングを確認
    if name in COMPANY_NAME_MAPPING:
        return COMPANY_NAME_MAPPING[name]
    
    # 部分一致でマッピングを探す
    for key, value in COMPANY_NAME_MAPPING.items():
        if key in name or name in key:
            return value
    
    return name

def normalize_price(price_str):
    """価格文字列を正規化（数値のみを抽出）"""
    if not price_str:
        return ''
    
    # 数値を抽出
    price_match = re.search(r'(\d{1,4}(?:[,，]\d{3})*(?:\.\d+)?)', str(price_str))
    if price_match:
        price_value = price_match.group(1).replace(',', '').replace('，', '')
        return price_value
    return ''

def apply_price_corrections_single(result, correction):
    """単一の結果に価格修正を適用（scrape_18_companies_to_excel.pyと同じロジック）"""
    if not isinstance(result, dict):
        print(f"ERROR: result is not a dict: {type(result)}")
        return result
    
    if not isinstance(correction, dict):
        print(f"ERROR: correction is not a dict: {type(correction)}")
        return result
    
    prices_raw = result.get('prices', {})
    if not isinstance(prices_raw, dict):
        print(f"WARNING: prices is not a dict: {type(prices_raw)}, value={prices_raw}")
        prices = {}
    else:
        prices = prices_raw.copy()
    
    # remove処理
    if 'remove' in correction and isinstance(correction['remove'], list):
        for material in correction['remove']:
            materials_to_remove = []
            for material_key in prices.keys():
                if material in material_key or material_key in material:
                    materials_to_remove.append(material_key)
            
            for material_key in materials_to_remove:
                if material_key in prices:
                    del prices[material_key]
    
    # add処理
    if 'add' in correction and isinstance(correction['add'], list):
        for item in correction['add']:
            if isinstance(item, dict) and 'material' in item and 'price' in item:
                material_name = item['material']
                price_value = item['price']
                price_normalized = normalize_price(price_value)
                if price_normalized:
                    prices[material_name] = price_normalized
    
    # modify処理
    if 'modify' in correction and isinstance(correction['modify'], list):
        for item in correction['modify']:
            if not isinstance(item, dict):
                continue
            old_material = item.get('material')
            if not old_material:
                continue
            new_price = item.get('price')
            new_material = item.get('material_new', old_material)
            
            # priceが存在する場合のみ正規化
            price_normalized = None
            if new_price:
                try:
                    price_normalized = normalize_price(new_price)
                except Exception as e:
                    print(f"WARNING: Failed to normalize price '{new_price}': {e}")
                    price_normalized = None
            
            # 材料名の部分一致で検索
            matched_material = None
            for material_key in prices.keys():
                if old_material in material_key or material_key in old_material:
                    matched_material = material_key
                    break
            
            if matched_material:
                if new_material != old_material:
                    # 材料名を変更（価格は既存の価格を維持）
                    prices[new_material] = prices[matched_material]
                    del prices[matched_material]
                    # 価格が指定されている場合は上書き
                    if price_normalized:
                        prices[new_material] = price_normalized
                else:
                    # 材料名は同じで価格のみ変更
                    if price_normalized:
                        prices[matched_material] = price_normalized
    
    result['prices'] = prices
    return result

# データベース初期化（Flask 2.3以降対応）
with app.app_context():
    db.create_all()

# 本番環境用の設定
if __name__ == '__main__':
    # 開発環境でのみ実行
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(debug=debug, host='0.0.0.0', port=port)






