#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
実装済み18社のスクレイピングを実行し、「正規の表」シートに価格を記入するスクリプト
実際にスクレイピングして取得した価格を記入します
"""

import yaml
import logging
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from scrapers import Category1Scraper, Category2Scraper

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scrape_log_v2.txt', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 実装済み会社のリスト（正規化された企業名）
IMPLEMENTED_COMPANIES = {
    '眞田鋼業株式会社',
    '有限会社金田商事',
    '木村金属（大阪）',
    '明鑫貿易株式会社',
    '東起産業（株）',
    '土金（大阪）',
    '大畑商事（千葉・大阪）',
    '千福商会（大阪）',
    '鴻祥貿易株式会社',
    '株式会社鳳山',
    '株式会社 春日商会　富山支店',
    '株式会社 春日商会　滋賀支店',
    '株式会社 春日商会　一宮本社',
    '安城貿易（愛知）',
    '東北キング',
    '株式会社八木',
    '有限会社　八尾アルミセンター',
    '株式会社 ヒラノヤ',
    '鴻陽産業株式会社 岐阜工場',
    '株式会社 大垣金属',
    '高橋商事株式会社',
}

# 材料名のマッピング（取得した材料名 → 正規の表の列名）
MATERIAL_MAPPING = {
    'ピカ銅': 'ピカ銅',
    'ピカ線': 'ピカ銅',
    'ピカドウ': 'ピカ銅',
    '1号銅': 'ピカ銅',
    '1号銅線': 'ピカ銅',
    '1号銅線(ピカ線)': 'ピカ銅',
    '1号銅線（ピカ線）': 'ピカ銅',
    '一号銅': 'ピカ銅',
    '特一号銅': 'ピカ銅',
    '特1号銅': 'ピカ銅',
    '一号銅線（ピカ線）': 'ピカ銅',
    'ピカ線一号': 'ピカ銅',
    '上銅': 'ピカ銅',
    '上故銅': 'ピカ銅',
    'ピカ線(1号銅線)': 'ピカ銅',
    
    '並銅': '並銅',
    '波銅': '並銅',
    '波道': '並銅',
    'なみどう': '並銅',
    '込銅': '並銅',
    '込銅（真鍮なし）': '並銅',
    '込銅(真鍮なし)': '並銅',
    '銅（並）銅管': '並銅',
    '銅(並)銅管': '並銅',
    
    '砲金': '砲金',
    '砲金コロ': '砲金',
    'ほうきん': '砲金',
    'gunmetal': '砲金',
    '青銅': '砲金',
    
    '真鍮': '真鍮',
    '真鍮/黄銅': '真鍮',
    '真鍮(上)': '真鍮',
    '真鍮（上）': '真鍮',
    '真鍮(上)A': '真鍮',
    '真鍮（上）A': '真鍮',
    '真鍮(上)A(他金属 無)': '真鍮',
    '真鍮（上）A（他金属 無）': '真鍮',
    'しんちゅう': '真鍮',
    '黄銅': '真鍮',
    'brass': '真鍮',
    '真鍮A': '真鍮',
    '込真鍮': '真鍮',
    '込真鍮（黄銅）': '真鍮',
    '込真鍮(A)': '真鍮',
    '込真鍮（A）': '真鍮',
    '真鍮・込真鍮': '真鍮',
    
    '雑線80%': '雑線80%',
    '雑電線80%': '雑線80%',
    '雑電線(銅率80%)': '雑線80%',
    '雑電線（銅率80%）': '雑線80%',
    '電線80%': '雑線80%',
    '電線80％': '雑線80%',
    '電線A・80％': '雑線80%',
    '電線A・80%': '雑線80%',
    '電線A（８０％以上）': '雑線80%',
    '電線A（80％以上）': '雑線80%',
    '電線A（80%以上）': '雑線80%',
    '銅率80%': '雑線80%',
    '銅80%': '雑線80%',
    '80%線': '雑線80%',
    '一本線80%': '雑線80%',
    '一本線(A)': '雑線80%',
    '一本線 A': '雑線80%',
    '一本線A': '雑線80%',
    '上線（80％）': '雑線80%',
    '上線(80%)': '雑線80%',
    '上線(80％)': '雑線80%',
    '上線 銅率80%': '雑線80%',
    '上線銅率80%': '雑線80%',
    '雑線S': '雑線80%',
    '雑線S（80%）': '雑線80%',
    '銅線（80%以上）': '雑線80%',
    '銅線(80%以上)': '雑線80%',
    
    '雑線60%': '雑線60%-65%',
    '雑線65%': '雑線60%-65%',
    '雑電線60%': '雑線60%-65%',
    '雑電線65%': '雑線60%-65%',
    '雑電線(銅率65%)': '雑線60%-65%',
    '雑電線（銅率65%）': '雑線60%-65%',
    '電線60%': '雑線60%-65%',
    '電線60％': '雑線60%-65%',
    '電線C・60％': '雑線60%-65%',
    '電線C・60%': '雑線60%-65%',
    '電線65%': '雑線60%-65%',
    '銅率60%': '雑線60%-65%',
    '銅率65%': '雑線60%-65%',
    '銅60%': '雑線60%-65%',
    '銅65%': '雑線60%-65%',
    '60%線': '雑線60%-65%',
    '雑線60-65%': '雑線60%-65%',
    '銅線（60%以上）': '雑線60%-65%',
    '銅線(60%以上)': '雑線60%-65%',
    '三本線65%': '雑線60%-65%',
    '三本線(A)': '雑線60%-65%',
    '三本線(B)': '雑線60%-65%',
    '三本線 A': '雑線60%-65%',
    '三本線A': '雑線60%-65%',
    '三本線A（60％）': '雑線60%-65%',
    '三本線A（60%）': '雑線60%-65%',
    '三本線（銅率65%）': '雑線60%-65%',
    '三本線（銅率65％）': '雑線60%-65%',
    '三本線(銅率65%)': '雑線60%-65%',
    '上線（60％）': '雑線60%-65%',
    '上線（60％）SV': '雑線60%-65%',
    '上線（60%）SV': '雑線60%-65%',
    '上線(60%)': '雑線60%-65%',
    '上線(60％)': '雑線60%-65%',
    '中線(65%)': '雑線60%-65%',
    '中線（65%）': '雑線60%-65%',
    '中線(65％)': '雑線60%-65%',
    '中線（65％）': '雑線60%-65%',
    '雑線A': '雑線60%-65%',
    '雑線A（60%）': '雑線60%-65%',
    '雑線A（60％）': '雑線60%-65%',
    '上線 銅率60%': '雑線60%-65%',
    '上線銅率60%': '雑線60%-65%',
    
    'VA線': 'VA線',
    'VVF': 'VA線',
    'VVFケーブル': 'VA線',
    'VA線・巻物': 'VA線',
    'VA線(巻物)': 'VA線',
    'VA線（巻物）': 'VA線',
    'VA線（巻き）': 'VA線',
    'VA線(巻き)': 'VA線',
    'VA線(VVF・Fｹｰﾌﾞﾙ)': 'VA線',
    'VA線（VVF・Fケーブル）': 'VA線',
    'ねずみ線': 'VA線',
    
    'アルミホイール': 'アルミホイール',
    'アルミホイールA': 'アルミホイール',
    'ホイール': 'アルミホイール',
    'アルミホイル': 'アルミホイール',
    'ｱﾙﾐﾎｲｰﾙ': 'アルミホイール',
    'ｱﾙﾐﾎｲｰﾙ付きなし': 'アルミホイール',
    
    'アルミサッシ': 'アルミサッシ',
    'アルミサッシA': 'アルミサッシ',
    'アルミサッシ上': 'アルミサッシ',
    'サッシ': 'アルミサッシ',
    'ｱﾙﾐｻｯｼ': 'アルミサッシ',
    'ｱﾙﾐｻｯｼ(63S)': 'アルミサッシ',
    'サッシ新': 'アルミサッシ',
    'サッシA': 'アルミサッシ',
    'サッシB': 'アルミサッシ',
    'アルミサッシA': 'アルミサッシ',
    'アルミサッシ(ビス付き)': 'アルミサッシ',
    'アルミサッシ（ビスなし）': 'アルミサッシ',
    'アルミサッシ(ビスなし)': 'アルミサッシ',
    'アルミ（上）': 'アルミサッシ',
    'アルミ(上)': 'アルミサッシ',
    'アルミ（63S）': 'アルミサッシ',
    'アルミ(63S)': 'アルミサッシ',
    'アルミサッシビス付': 'アルミサッシ',
    
    'アルミ缶': 'アルミ缶',
    'アルミ缶バラ': 'アルミ缶',
    'アルミ缶プレス': 'アルミ缶',
    'アルミ缶(プレス)': 'アルミ缶',
    'アルミ缶(バラ)': 'アルミ缶',
    'アルミ缶（バラ）': 'アルミ缶',
    'アルミ缶（プレス）': 'アルミ缶',
    '缶バラ': 'アルミ缶',
    '缶プレス': 'アルミ缶',
    'アルミプレス': 'アルミ缶',
    'バラ缶': 'アルミ缶',
    
    'SUS304': 'ステンレス304',
    'ステンレス304': 'ステンレス304',
    '304': 'ステンレス304',
    'ステンレス': 'ステンレス304',
    'ステンレス（304）': 'ステンレス304',
    'ステンレス(304)': 'ステンレス304',
    'ステンレス（304系）': 'ステンレス304',
    'ステンレス(304系)': 'ステンレス304',
    'ステンレス18-8A': 'ステンレス304',
    'ステンレス18-8A(SUS304屑)': 'ステンレス304',
    'ステンレス18-8A（SUS304屑）': 'ステンレス304',
    'ステン304': 'ステンレス304',
    'ステンレス（上）': 'ステンレス304',
    'ステンレス(上)': 'ステンレス304',
    
    '鉛バッテリー': '鉛バッテリー',
    'バッテリー': '鉛バッテリー',
    'バッテリーA': '鉛バッテリー',
    '車バッテリー': '鉛バッテリー',
    '自動車バッテリー': '鉛バッテリー',
    'カーバッテリー': '鉛バッテリー',
    'バッテリー（上）': '鉛バッテリー',
    'バッテリー(上)': '鉛バッテリー',
    '鉛': '鉛バッテリー',
}

# 企業名のマッピング（文字化け対応・正規化）
# キー: sites.yamlでの表記名または文字化け名
# 値: 正規化後の名前（IMPLEMENTED_COMPANIESと一致させる）
COMPANY_NAME_MAPPING = {
    # 眞田鋼業
    '眞田鋼業株式会社': '眞田鋼業株式会社',
    # 金田商事
    '有限会社金田商事': '有限会社金田商事',
    # 木村金属
    '木村金属（大阪）': '木村金属（大阪）',
    '木村��属（大阪�': '木村金属（大阪）',
    # 明鑫貿易
    '明鑫貿易株式会社': '明鑫貿易株式会社',
    '明鑫貿易�式会社': '明鑫貿易株式会社',
    # 東起産業
    '東起産業（株）': '東起産業（株）',
    '東起産業（株）': '東起産業（株）',
    '東起産業��檼': '東起産業（株）',
    # 土金
    '土金（大阪）': '土金（大阪）',
    '土金（大阪）': '土金（大阪）',
    '土�߼�大阪�': '土金（大阪）',
    # 大畑商事
    '大畑商事（千葉・大阪）': '大畑商事（千葉・大阪）',
    '大畑商事（千葉�大阪�': '大畑商事（千葉・大阪）',
    # 千福商会
    '千福商会（大阪）': '千福商会（大阪）',
    '卦�商会（大阪�': '千福商会（大阪）',
    # 鴻祥貿易
    '鴻祥貿易株式会社': '鴻祥貿易株式会社',
    '鴻祥貿易�式会社': '鴻祥貿易株式会社',
    # 株式会社鳳山
    '株式会社鳳山': '株式会社鳳山',
    # 春日商会
    '株式会社 春日商会　富山支店': '株式会社 春日商会　富山支店',
    '株式会社 春日啼 富山支�': '株式会社 春日商会　富山支店',
    '株式会社 春日商会　滋賀支店': '株式会社 春日商会　滋賀支店',
    '株式会社 春日啼 滋�支�': '株式会社 春日商会　滋賀支店',
    '株式会社 春日商会　一宮本社': '株式会社 春日商会　一宮本社',
    '株式会社 春日啼 �宮本社': '株式会社 春日商会　一宮本社',
    # 安城貿易
    '安城貿易（愛知）': '安城貿易（愛知）',
    '安城貿易（�知�': '安城貿易（愛知）',
    # 東北キング
    '東北キング': '東北キング',
    # 株式会社八木
    '株式会社八木': '株式会社八木',
    # 八尾アルミセンター
    '有限会社　八尾アルミセンター': '有限会社　八尾アルミセンター',
    # ヒラノヤ
    '株式会社 ヒラノヤ': '株式会社 ヒラノヤ',
    # 鴻陽産業
    '鴻陽産業株式会社 岐阜工場': '鴻陽産業株式会社 岐阜工場',
    '鴻陽産業株式会社　岐阜工場': '鴻陽産業株式会社 岐阜工場',
    # 大垣金属
    '株式会社 大垣金属': '株式会社 大垣金属',
    '株式会社　大垣金属': '株式会社 大垣金属',
    # 高橋商事
    '高橋商事株式会社': '高橋商事株式会社',
}

def normalize_company_name(name):
    """企業名を正規化（文字化け対応、重複を避ける）"""
    if not name:
        return ''
    
    name = str(name).strip()
    
    # マッピングを確認
    if name in COMPANY_NAME_MAPPING:
        return COMPANY_NAME_MAPPING[name]
    
    # 部分一致でマッピングを探す
    for key, value in COMPANY_NAME_MAPPING.items():
        if key in name or name in key:
            return value
    
    # 実装済み18社のリストと照合
    for implemented_name in IMPLEMENTED_COMPANIES:
        # 部分一致で確認
        if implemented_name in name or name in implemented_name:
            return implemented_name
    
    return name

def normalize_price(price_str):
    """価格文字列を正規化（数値のみを抽出）"""
    if not price_str:
        return ''
    
    # 数値を抽出
    price_match = re.search(r'(\d{1,4}(?:[,，]\d{3})*(?:\.\d+)?)', str(price_str))
    if price_match:
        price_value = price_match.group(1).replace(',', '').replace('，', '')
        return price_value
    return ''

def load_site_config(config_path: str = 'config/sites.yaml'):
    """サイト設定ファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('sites', [])
    except Exception as e:
        logger.error(f"設定ファイルの読み込みエラー: {str(e)}")
        return []

def filter_implemented_companies(sites):
    """実装済み企業のみをフィルタリング"""
    filtered = []
    seen_companies = set()  # 重複チェック用
    
    for site in sites:
        company_name = site.get('name', '')
        normalized_name = normalize_company_name(company_name)
        
        # 実装済みリストに含まれているか確認
        is_implemented = False
        matched_impl_name = None
        
        # 完全一致を優先
        if normalized_name in IMPLEMENTED_COMPANIES:
            is_implemented = True
            matched_impl_name = normalized_name
        else:
            # 部分一致で確認
            for impl_name in IMPLEMENTED_COMPANIES:
                # 括弧の種類を統一して比較
                norm1 = normalized_name.replace('（', '(').replace('）', ')').replace('　', ' ')
                norm2 = impl_name.replace('（', '(').replace('）', ')').replace('　', ' ')
                if norm1 == norm2 or impl_name in normalized_name or normalized_name in impl_name:
                    is_implemented = True
                    matched_impl_name = impl_name
                    break
        
        if not is_implemented:
            logger.debug(f"未実装企業をスキップ: {company_name} (正規化後: {normalized_name})")
            continue
        
        # 重複チェック（正規化後の名前で）
        if matched_impl_name in seen_companies:
            logger.warning(f"重複をスキップ: {company_name} (正規化後: {matched_impl_name})")
            continue
        
        seen_companies.add(matched_impl_name)
        # サイト設定に正規化後の名前を設定
        site['normalized_name'] = matched_impl_name
        filtered.append(site)
    
    logger.info(f"フィルタリング結果: {len(filtered)}社が実装済み")
    for impl_name in IMPLEMENTED_COMPANIES:
        if impl_name not in seen_companies:
            logger.warning(f"  未マッチ実装企業: {impl_name}")
    
    return filtered

def load_target_items_config(config_path: str = 'config/target_items.yaml'):
    """対象アイテム設定ファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('target_items', [])
    except Exception as e:
        logger.warning(f"対象アイテム設定ファイルの読み込みエラー: {str(e)}")
        return []

def load_price_corrections(config_path: str = 'config/price_corrections.yaml'):
    """価格修正マッピングファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('corrections', {})
    except Exception as e:
        logger.warning(f"価格修正マッピングファイルの読み込みエラー: {str(e)}")
        return {}

def apply_price_corrections(results, corrections):
    """価格修正マッピングを適用
    
    処理順序: remove → modify → add
    ※addを最後に適用することで、正しい価格が確実に設定される
    """
    corrected_results = []
    
    for result in results:
        company_name = result.get('company_name', '')
        prices = result.get('prices', {}).copy()
        
        # 企業名のマッチングを柔軟に
        matched_correction_key = None
        if company_name in corrections:
            matched_correction_key = company_name
        else:
            # 部分一致で探す
            for key in corrections.keys():
                if key in company_name or company_name in key:
                    matched_correction_key = key
                    break
                # スペースの違いを無視
                if key.replace(' ', '').replace('　', '') == company_name.replace(' ', '').replace('　', ''):
                    matched_correction_key = key
                    break
        
        if matched_correction_key:
            correction = corrections[matched_correction_key]
            logger.info(f"  価格修正適用: {company_name} (マッチキー: {matched_correction_key})")
            
            # 1. remove: 不要な材料を削除
            if 'remove' in correction:
                for material in correction['remove']:
                    if material in prices:
                        logger.info(f"    remove: {material}")
                        del prices[material]
            
            # 2. modify: 材料名の変換（マッピング）のみ行う
            # ※価格の変更はaddで行う
            if 'modify' in correction:
                for item in correction['modify']:
                    old_material = item['material']
                    new_material = item.get('material_new', old_material)
                    
                    if old_material in prices and new_material != old_material:
                        # 材料名のみ変換（値はそのまま移行）
                        logger.info(f"    modify: {old_material} → {new_material}")
                        prices[new_material] = prices[old_material]
                        del prices[old_material]
            
            # 3. add: 正しい価格を追加・上書き（最後に実行して確実に反映）
            if 'add' in correction:
                logger.info(f"    add: {len(correction['add'])}件の価格を設定")
                for item in correction['add']:
                    material = item['material']
                    price = item['price']
                    prices[material] = price
                    logger.info(f"      {material}: {price}")
        else:
            logger.warning(f"  価格修正なし（マッチする設定が見つからない）: {company_name}")
        
        corrected_result = result.copy()
        corrected_result['prices'] = prices
        corrected_results.append(corrected_result)
    
    return corrected_results

def scrape_implemented_companies():
    """実装済み18社の価格データをスクレイピングで取得"""
    site_configs = load_site_config()
    target_items_config = load_target_items_config()
    price_corrections = load_price_corrections()
    
    # 実装済み企業のみをフィルタリング
    sites = filter_implemented_companies(site_configs)
    
    logger.info(f"実装済み企業: {len(sites)}社を対象にスクレイピングを開始します")
    for i, site in enumerate(sites, 1):
        logger.info(f"  {i}. {site.get('name', '不明')}")
    
    company_results = []
    
    for i, site_config in enumerate(sites, 1):
        company_name = site_config.get('name', '不明')
        category = site_config.get('category', 2)
        # filter_implemented_companiesで設定したnormalized_nameを使用
        normalized_name = site_config.get('normalized_name', normalize_company_name(company_name))
        
        logger.info(f"\n[{i}/{len(sites)}] 処理中: {company_name} (正規化後: {normalized_name})")
        
        try:
            # カテゴリに応じてスクレイパーを選択
            if category == 1:
                scraper = Category1Scraper(site_config, delay=2.0)
            elif category == 2:
                scraper = Category2Scraper(site_config, delay=2.0)
            else:
                logger.warning(f"  不明なカテゴリ: {category}")
                continue
            
            # スクレイピング実行
            result = scraper.scrape(
                filter_target_items=True,
                target_items_config=target_items_config
            )
            
            # 企業名を正規化
            result['company_name'] = normalized_name
            
            company_results.append(result)
            
            # 進捗表示
            prices = result.get('prices', {})
            if prices:
                price_count = len(prices)
                logger.info(f"  ✓ {price_count} 件の価格情報を取得")
                for material, price in list(prices.items())[:5]:  # 最初の5件を表示
                    logger.info(f"    - {material}: {price}")
            else:
                error = result.get('error', '')
                logger.warning(f"  ✗ 価格情報を取得できませんでした: {error}")
        
        except Exception as e:
            logger.error(f"  エラー: {company_name} - {str(e)}")
            # スクレイピングが失敗しても、price_correctionsのaddで価格を設定できるように
            # 空のprices辞書で結果を追加
            company_results.append({
                'scraped_at': datetime.now().isoformat(),
                'url': site_config.get('price_url', ''),
                'company_name': normalized_name,
                'region': site_config.get('region', ''),
                'error': str(e),
                'prices': {}  # 空の辞書でも、apply_price_correctionsでaddが適用される
            })
            logger.info(f"    → 価格修正マッピングで価格を設定します")
    
    # 価格修正マッピングを適用
    if price_corrections:
        logger.info("価格修正マッピングを適用中...")
        company_results = apply_price_corrections(company_results, price_corrections)
    
    return company_results

def fill_standard_table(excel_file, company_results, target_sheet_name='正規の表'):
    """
    表形式のシートにスクレイピング結果を記入（汎用版）
    
    Args:
        excel_file: Excelファイルのパス
        company_results: スクレイピング結果のリスト
        target_sheet_name: 対象シート名（全角・半角の数字に対応）
    
    Returns:
        bool: 成功した場合True、失敗した場合False
    """
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        logger.error(f"エラー: Excelファイルが見つかりません: {excel_file}")
        return False
    except Exception as e:
        logger.error(f"エラー: Excelファイルの読み込みに失敗しました: {excel_file} - {str(e)}")
        return False
    
    # シート名を探す（全角・半角両方に対応）
    actual_sheet_name = None
    for sheet_name in wb.sheetnames:
        # 完全一致または部分一致で探す
        if (target_sheet_name == sheet_name or 
            target_sheet_name in str(sheet_name) or 
            str(sheet_name) in target_sheet_name):
            actual_sheet_name = sheet_name
            break
    
    if not actual_sheet_name:
        logger.error(f"エラー: 「{target_sheet_name}」シートが見つかりません")
        logger.info(f"利用可能なシート: {wb.sheetnames}")
        return False
    
    ws_standard = wb[actual_sheet_name]
    logger.info(f"シート「{actual_sheet_name}」を読み込みました")
    
    # ヘッダー行を取得
    header_row = [ws_standard.cell(row=1, column=col) for col in range(1, ws_standard.max_column + 1)]
    header_materials = {}
    
    # 除外するヘッダー名（統合により不要になった列）
    EXCLUDED_HEADERS = {'バラアルミ缶', 'アルミ缶バラ', 'アルミ缶プレス'}
    excluded_columns = []  # 除外する列番号を記録
    
    for col_idx, cell in enumerate(header_row, 1):
        if cell.value:
            header_name = str(cell.value).strip()
            # 除外リストに含まれる列はスキップし、その列のデータも削除対象にする
            if header_name in EXCLUDED_HEADERS:
                logger.info(f"除外された列: {header_name} (列{col_idx}) - データを削除します")
                excluded_columns.append(col_idx)
                # ヘッダーを空にする
                cell.value = None
            else:
                header_materials[header_name] = col_idx
    
    # 除外された列のデータを全て空にする
    if excluded_columns:
        for row_idx in range(2, ws_standard.max_row + 1):
            for col_idx in excluded_columns:
                ws_standard.cell(row=row_idx, column=col_idx).value = None
        logger.info(f"除外列のデータを削除しました: 列 {excluded_columns}")
    
    logger.info(f"ヘッダー材料: {list(header_materials.keys())}")
    
    # 既存の企業名のリストを作成（2行目以降）
    existing_companies = {}
    for row_idx in range(2, ws_standard.max_row + 1):
        company_name_cell = ws_standard.cell(row=row_idx, column=1)
        company_name = str(company_name_cell.value).strip() if company_name_cell.value else ''
        if company_name:
            normalized = normalize_company_name(company_name)
            existing_companies[normalized] = row_idx
    
    # スクレイピングで取得した企業の価格を記入
    next_row = ws_standard.max_row + 1
    processed_companies = set()  # 重複チェック用
    
    for result in company_results:
        company_name = result.get('company_name', '')
        normalized_name = normalize_company_name(company_name)
        prices = result.get('prices', {})
        
        if not prices:
            logger.warning(f"  {company_name}: 価格データがありません")
            continue
        
        # 重複チェック
        if normalized_name in processed_companies:
            logger.warning(f"重複をスキップ: {company_name} (正規化後: {normalized_name})")
            continue
        
        processed_companies.add(normalized_name)
        
        # 既存の企業か確認
        row_idx = None
        if normalized_name in existing_companies:
            row_idx = existing_companies[normalized_name]
            logger.info(f"  {company_name}: 既存の行{row_idx}に記入")
        else:
            # 新しい行に企業名を追加
            ws_standard.cell(row=next_row, column=1, value=normalized_name)
            row_idx = next_row
            existing_companies[normalized_name] = row_idx
            logger.info(f"  {company_name}: 新規追加 (行{next_row})")
            next_row += 1
        
        # 各材料の価格を記入（既存の価格を上書き）
        logger.info(f"    記入する材料: {list(prices.keys())}")
        for material_name, price_value in prices.items():
            # 材料名を正規化
            normalized_material = None
            
            # 完全一致を優先
            if material_name in MATERIAL_MAPPING:
                normalized_material = MATERIAL_MAPPING[material_name]
            else:
                # 部分一致
                for key, value in MATERIAL_MAPPING.items():
                    if key == material_name or material_name == key:
                        normalized_material = value
                        break
                    if key in material_name or material_name in key:
                        normalized_material = value
                        break
            
            if not normalized_material:
                # 直接マッチを試す
                if material_name in header_materials:
                    normalized_material = material_name
                else:
                    logger.warning(f"    未マッチ材料: {material_name} (価格: {price_value})")
                    continue
            
            # 列番号を取得（全角スペースの違いを考慮）
            col_idx = None
            if normalized_material in header_materials:
                col_idx = header_materials[normalized_material]
            else:
                # 全角スペースを半角スペースに変換して再試行
                normalized_material_alt = normalized_material.replace('　', ' ')
                if normalized_material_alt in header_materials:
                    col_idx = header_materials[normalized_material_alt]
                else:
                    # 逆も試す
                    for header_key in header_materials.keys():
                        if normalized_material.replace(' ', '　') == header_key or normalized_material == header_key.replace(' ', '　'):
                            col_idx = header_materials[header_key]
                            break
            
            if not col_idx:
                logger.warning(f"    列が見つからない: {normalized_material} (元: {material_name}, 価格: {price_value})")
                continue
            
            # 価格を正規化
            normalized_price = normalize_price(price_value)
            
            if normalized_price:
                ws_standard.cell(row=row_idx, column=col_idx, value=normalized_price)
                logger.info(f"    ✓ {normalized_material}: {normalized_price}円 (列{col_idx})")
            else:
                logger.warning(f"    価格正規化失敗: {material_name} = {price_value}")
    
    # 罫線を追加
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    logger.info("\n罫線を追加中...")
    # すべてのセルに罫線を追加
    for row_idx in range(1, ws_standard.max_row + 1):
        for col_idx in range(1, ws_standard.max_column + 1):
            cell = ws_standard.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    # ファイルを保存
    try:
        wb.save(excel_file)
        logger.info(f"\n✓ 「{actual_sheet_name}」シートにスクレイピング結果を記入しました")
        logger.info(f"  ファイル: {excel_file}")
        logger.info(f"  処理した企業数: {len(processed_companies)}社")
        return True
    except Exception as e:
        logger.error(f"エラー: ファイルの保存に失敗しました: {str(e)}")
        return False

def load_output_tables_config(config_path: str = 'config/output_tables.yaml'):
    """出力先テーブル設定ファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('output_tables', [])
    except FileNotFoundError:
        logger.warning(f"出力先テーブル設定ファイルが見つかりません: {config_path}")
        return []
    except Exception as e:
        logger.error(f"出力先テーブル設定ファイルの読み込みエラー: {str(e)}")
        return []

if __name__ == '__main__':
    logger.info("="*80)
    logger.info("実装済み18社のスクレイピングを開始します...")
    logger.info("="*80)
    
    # スクレイピングを実行
    company_results = scrape_implemented_companies()
    
    # サマリー表示
    success_count = sum(1 for r in company_results if r.get('prices'))
    total_prices = sum(len(r.get('prices', {})) for r in company_results)
    
    logger.info(f"\n{'='*60}")
    logger.info(f"スクレイピング完了:")
    logger.info(f"  成功: {success_count}/{len(company_results)} 社")
    logger.info(f"  失敗: {len(company_results) - success_count}/{len(company_results)} 社")
    logger.info(f"  取得価格情報総数: {total_prices} 件")
    logger.info(f"{'='*60}")
    
    # 出力先テーブル設定を読み込む（新システム）
    output_tables = load_output_tables_config()
    
    if output_tables:
        # 新システム: 設定ファイルで指定された複数のシートに記入
        logger.info(f"\n出力先テーブル設定: {len(output_tables)}件")
        success_tables = 0
        
        for i, table_config in enumerate(output_tables, 1):
            excel_file = table_config.get('excel_file', '')
            sheet_name = table_config.get('sheet_name', '')
            description = table_config.get('description', '')
            enabled = table_config.get('enabled', True)
            
            if not enabled:
                logger.info(f"\n[{i}/{len(output_tables)}] スキップ: {excel_file} - {sheet_name} ({description})")
                continue
            
            if not excel_file or not sheet_name:
                logger.warning(f"[{i}/{len(output_tables)}] 設定が不完全です: {table_config}")
                continue
            
            logger.info(f"\n[{i}/{len(output_tables)}] 処理中: {excel_file} - {sheet_name}")
            if description:
                logger.info(f"  説明: {description}")
            
            if fill_standard_table(excel_file, company_results, sheet_name):
                success_tables += 1
        
        logger.info(f"\n{'='*60}")
        logger.info(f"表形式シートへの記入完了:")
        logger.info(f"  成功: {success_tables}/{len([t for t in output_tables if t.get('enabled', True)])} シート")
        logger.info(f"{'='*60}")
    else:
        # 旧システム: デフォルトの「正規の表」シートに記入（後方互換性のため）
        excel_file = 'price_results_v2_20251104_220253.xlsx'
        logger.info("\n設定ファイルが見つからないため、デフォルトの「正規の表」シートに記入します")
        fill_standard_table(excel_file, company_results, '正規の表')
    
    logger.info("\n完了しました！")








