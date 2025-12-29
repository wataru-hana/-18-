#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WordPressの相場情報ページから表データを抽出してExcelに出力するスクリプト
今後の記録用として相場情報を保存するために使用します。

使用方法:
    python save_market_prices_to_excel.py
    
    ※URLを変更する場合:
    python save_market_prices_to_excel.py --url "https://hitetsunavi.jp/target-page/"
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import argparse
import sys

def fetch_and_export(target_url):
    print(f"アクセス中: {target_url} ...")
    
    try:
        # ヘッダーを設定してブラウザからのアクセスに見せる
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(target_url, headers=headers)
        response.raise_for_status()
        response.encoding = response.apparent_encoding
    except Exception as e:
        print(f"エラー: URLへのアクセスに失敗しました。\n詳細: {e}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    
    # タイトル取得
    page_title = soup.title.string if soup.title else "No Title"
    print(f"ページタイトル: {page_title}")

    # テーブルを抽出
    # wp-block-tableクラスを持つfigureの中のtable、または単純なtableタグを探す
    tables = soup.find_all('table')
    
    if not tables:
        print("エラー: ページ内に表（テーブル）が見つかりませんでした。")
        return

    print(f"{len(tables)} 個の表が見つかりました。Excel作成を開始します...")

    # Excelファイル作成
    wb = Workbook()
    
    # サマリーシート作成
    ws_summary = wb.active
    ws_summary.title = "出力情報"
    ws_summary.append(["項目", "内容"])
    ws_summary.append(["取得日時", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_summary.append(["取得URL", target_url])
    ws_summary.append(["ページタイトル", page_title])
    ws_summary.append(["検出テーブル数", len(tables)])
    
    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i, table in enumerate(tables):
        sheet_name = f"表_{i+1}"
        ws = wb.create_sheet(title=sheet_name)
        
        rows = []
        
        # テーブルデータの抽出
        # theadがある場合
        thead = table.find('thead')
        if thead:
            header_row = []
            for th in thead.find_all(['th', 'td']):
                header_row.append(th.get_text(strip=True))
            rows.append(header_row)
            
            # tbody
            tbody = table.find('tbody')
            if tbody:
                for tr in tbody.find_all('tr'):
                    row_data = []
                    for td in tr.find_all(['td', 'th']):
                        row_data.append(td.get_text(strip=True))
                    rows.append(row_data)
        else:
            # theadがない場合はすべてのtrを処理
            for tr in table.find_all('tr'):
                row_data = []
                for td in tr.find_all(['td', 'th']):
                    row_data.append(td.get_text(strip=True))
                rows.append(row_data)
        
        # Excelに書き込み
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                
                # ヘッダー行（1行目）のスタイル適用
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 列幅自動調整（簡易版）
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        # 全角文字が含まれる場合は幅を広めに取る
                        for char in str(cell.value):
                            if ord(char) > 255:
                                cell_len += 1
                        if cell_len > max_length:
                            max_length = cell_len
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) # 最大幅を50に制限
            ws.column_dimensions[column_letter].width = adjusted_width

    # 保存
    filename = f"相場情報記録_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(os.getcwd(), filename)
    
    try:
        wb.save(output_path)
        print("-" * 50)
        print(f"完了しました。以下のファイルに保存されました:")
        print(f"{output_path}")
        print("-" * 50)
    except Exception as e:
        print(f"エラー: ファイルの保存に失敗しました。\n詳細: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='WordPressのページから表データを抽出してExcelに保存します。')
    parser.add_argument('--url', default='https://hitetsunavi.jp/scrap-metal-prices/', 
                        help='スクレイピング対象のURL (デフォルト: https://hitetsunavi.jp/scrap-metal-prices/)')
    
    args = parser.parse_args()
    
    print("スクリプトを開始します...")
    fetch_and_export(args.url)







