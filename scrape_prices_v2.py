#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
非鉄金属スクラップ価格自動取得スクリプト（改良版）
設定ファイルベースで各サイトの価格情報を取得します
"""

import yaml
import json
import csv
import logging
from datetime import datetime
from typing import List, Dict
from pathlib import Path

# ログ設定（先に設定）
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scrape_log_v2.txt', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Excel出力ライブラリのインポート
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxlがインストールされていません。Excel出力機能は使用できません。")

from scrapers import Category1Scraper, Category2Scraper


def load_site_config(config_path: str = 'config/sites.yaml') -> List[Dict]:
    """
    サイト設定ファイルを読み込む
    
    Args:
        config_path: 設定ファイルのパス
        
    Returns:
        サイト設定のリスト
    """
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('sites', [])
    except FileNotFoundError:
        logger.error(f"設定ファイルが見つかりません: {config_path}")
        return []
    except Exception as e:
        logger.error(f"設定ファイルの読み込みエラー: {str(e)}")
        return []


def load_target_items_config(config_path: str = 'config/target_items.yaml') -> List[Dict]:
    """
    対象アイテム設定ファイルを読み込む
    
    Args:
        config_path: 設定ファイルのパス
        
    Returns:
        対象アイテム設定のリスト
    """
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('target_items', [])
    except FileNotFoundError:
        logger.warning(f"対象アイテム設定ファイルが見つかりません: {config_path}（全アイテムを抽出します）")
        return []
    except Exception as e:
        logger.warning(f"対象アイテム設定ファイルの読み込みエラー: {str(e)}（全アイテムを抽出します）")
        return []


def load_price_corrections(config_path: str = 'config/price_corrections.yaml') -> Dict:
    """
    価格修正マッピングファイルを読み込む
    
    Args:
        config_path: 設定ファイルのパス
        
    Returns:
        修正マッピングの辞書
    """
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('corrections', {})
    except FileNotFoundError:
        logger.warning(f"価格修正マッピングファイルが見つかりません: {config_path}")
        return {}
    except Exception as e:
        logger.warning(f"価格修正マッピングファイルの読み込みエラー: {str(e)}")
        return {}


def apply_price_corrections(results: List[Dict], corrections: Dict) -> List[Dict]:
    """
    価格修正マッピングを適用
    
    Args:
        results: スクレイピング結果のリスト
        corrections: 修正マッピングの辞書
        
    Returns:
        修正後の結果リスト
    """
    corrected_results = []
    
    for result in results:
        company_name = result.get('company_name', '')
        prices = result.get('prices', {}).copy()
        
        # 該当する修正マッピングがあるか確認
        if company_name in corrections:
            correction = corrections[company_name]
            
            # 削除
            if 'remove' in correction:
                for material in correction['remove']:
                    if material in prices:
                        del prices[material]
            
            # 追加
            if 'add' in correction:
                for item in correction['add']:
                    prices[item['material']] = item['price']
            
            # 修正
            if 'modify' in correction:
                for item in correction['modify']:
                    old_material = item['material']
                    new_price = item['price']
                    new_material = item.get('material_new', old_material)
                    
                    if old_material in prices:
                        # 材料名が変更される場合
                        if new_material != old_material:
                            # 価格が文字列（材料名など）の場合は元の価格を保持
                            if new_price == new_material or new_price == old_material:
                                # 価格が材料名と同じ場合は、元の価格を保持
                                prices[new_material] = prices[old_material]
                            else:
                                prices[new_material] = new_price
                            del prices[old_material]
                        else:
                            prices[old_material] = new_price
        
        # 修正後の結果を作成
        corrected_result = result.copy()
        corrected_result['prices'] = prices
        corrected_results.append(corrected_result)
    
    return corrected_results


def save_results(results: List[Dict], output_format: str = 'json'):
    """
    結果をファイルに保存
    
    Args:
        results: スクレイピング結果のリスト
        output_format: 出力形式 ('json', 'csv', または 'excel')
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if output_format == 'json':
        output_file = f'price_results_v2_{timestamp}.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        logger.info(f"結果を {output_file} に保存しました")
    
    elif output_format == 'csv':
        output_file = f'price_results_v2_{timestamp}.csv'
        if results:
            rows = []
            for result in results:
                company_name = result.get('company_name', '')
                url = result.get('url', '')
                region = result.get('region', '')
                scraped_at = result.get('scraped_at', '')
                
                prices = result.get('prices', {})
                if prices:
                    for material, price in prices.items():
                        rows.append({
                            '会社名': company_name,
                            'URL': url,
                            '地域': region,
                            '材料名': material,
                            '価格': price,
                            '取得日時': scraped_at
                        })
                else:
                    error = result.get('error', '')
                    rows.append({
                        '会社名': company_name,
                        'URL': url,
                        '地域': region,
                        '材料名': '',
                        '価格': '',
                        '取得日時': scraped_at,
                        'エラー': error
                    })
            
            if rows:
                fieldnames = ['会社名', 'URL', '地域', '材料名', '価格', '取得日時', 'エラー']
                with open(output_file, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
                logger.info(f"結果を {output_file} に保存しました")
    
    elif output_format == 'excel':
        if not EXCEL_AVAILABLE:
            logger.error("Excel出力にはopenpyxlが必要です。'pip install openpyxl'を実行してください。")
            return
        
        # 既存のExcelファイルが存在する場合は、そのファイルに別シートとして追加
        existing_file = 'price_results_v2_20251104_220253.xlsx'
        if Path(existing_file).exists():
            from openpyxl import load_workbook
            wb = load_workbook(existing_file)
            # 新しいシート名を生成（タイムスタンプ付き）
            base_sheet_name = f"価格情報_{timestamp}"
            sheet_name = base_sheet_name
            
            # シート名の重複チェック（同じ名前のシートが存在する場合は番号を追加）
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_sheet_name}_{counter}"
                counter += 1
            
            ws = wb.create_sheet(title=sheet_name)
            output_file = existing_file
            logger.info(f"既存のExcelファイル '{existing_file}' に新しいシート '{sheet_name}' を追加します")
        else:
            output_file = f'price_results_v2_{timestamp}.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = "価格情報"
            logger.info(f"新しいExcelファイル '{output_file}' を作成します")
        
        # ヘッダーの設定
        headers = ['会社名', 'URL', '地域', '材料名', '価格', '取得日時', 'エラー']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行を書き込み
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # データ行を書き込み
        row_idx = 2
        for result in results:
            company_name = result.get('company_name', '')
            url = result.get('url', '')
            region = result.get('region', '')
            scraped_at = result.get('scraped_at', '')
            
            prices = result.get('prices', {})
            if prices:
                for material, price in prices.items():
                    ws.cell(row=row_idx, column=1, value=company_name).border = border
                    ws.cell(row=row_idx, column=2, value=url).border = border
                    ws.cell(row=row_idx, column=3, value=region).border = border
                    ws.cell(row=row_idx, column=4, value=material).border = border
                    ws.cell(row=row_idx, column=5, value=price).border = border
                    ws.cell(row=row_idx, column=6, value=scraped_at).border = border
                    ws.cell(row=row_idx, column=7, value='').border = border
                    row_idx += 1
            else:
                error = result.get('error', '')
                ws.cell(row=row_idx, column=1, value=company_name).border = border
                ws.cell(row=row_idx, column=2, value=url).border = border
                ws.cell(row=row_idx, column=3, value=region).border = border
                ws.cell(row=row_idx, column=4, value='').border = border
                ws.cell(row=row_idx, column=5, value='').border = border
                ws.cell(row=row_idx, column=6, value=scraped_at).border = border
                ws.cell(row=row_idx, column=7, value=error).border = border
                row_idx += 1
        
        # 列幅の自動調整
        column_widths = {
            'A': 30,  # 会社名
            'B': 50,  # URL
            'C': 10,  # 地域
            'D': 25,  # 材料名
            'E': 20,  # 価格
            'F': 20,  # 取得日時
            'G': 30   # エラー
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # ヘッダー行の高さを設定
        ws.row_dimensions[1].height = 25
        
        # データ行の文字列折り返し設定
        for row in ws.iter_rows(min_row=2, max_row=row_idx-1):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        wb.save(output_file)
        # ログメッセージを出力
        if output_file == existing_file and 'sheet_name' in locals():
            logger.info(f"✓ 取得結果を {output_file} のシート '{sheet_name}' に保存しました")
        else:
            logger.info(f"✓ 結果を {output_file} に保存しました")


def main():
    """メイン処理"""
    # 設定ファイルを読み込む
    sites = load_site_config('config/sites.yaml')
    
    if not sites:
        logger.error("サイト設定の読み込みに失敗しました")
        return
    
    # 対象アイテム設定を読み込む
    target_items = load_target_items_config('config/target_items.yaml')
    filter_enabled = len(target_items) > 0
    
    if filter_enabled:
        logger.info(f"対象アイテムフィルタリング: 有効 ({len(target_items)}種類)")
        for item in target_items:
            logger.info(f"  - {item.get('name', '')}")
    else:
        logger.info("対象アイテムフィルタリング: 無効（全アイテムを抽出）")
    
    # 価格修正マッピングを読み込む
    price_corrections = load_price_corrections('config/price_corrections.yaml')
    if price_corrections:
        logger.info(f"価格修正マッピング: {len(price_corrections)} 社の修正を適用します")
    
    logger.info(f"スクレイピング開始: {len(sites)} 社")
    
    results = []
    
    for i, site_config in enumerate(sites, 1):
        company_name = site_config.get('name', '不明')
        category = site_config.get('category', 0)
        
        logger.info(f"[{i}/{len(sites)}] 処理中: {company_name} (カテゴリ{category})")
        
        try:
            # カテゴリに応じてスクレイパーを選択
            if category == 1:
                scraper = Category1Scraper(site_config, delay=2.0)
            elif category == 2:
                scraper = Category2Scraper(site_config, delay=2.0)
            else:
                logger.warning(f"  不明なカテゴリ: {category}")
                continue
            
            # スクレイピング実行
            result = scraper.scrape(
                filter_target_items=filter_enabled,
                target_items_config=target_items if filter_enabled else None
            )
            results.append(result)
            
            # 進捗表示
            prices = result.get('prices', {})
            if prices:
                price_count = len(prices)
                logger.info(f"  ✓ {price_count} 件の価格情報を取得")
            else:
                error = result.get('error', '')
                logger.warning(f"  ✗ 価格情報を取得できませんでした: {error}")
        
        except Exception as e:
            logger.error(f"  エラー: {company_name} - {str(e)}")
            results.append({
                'scraped_at': datetime.now().isoformat(),
                'url': site_config.get('price_url', ''),
                'company_name': company_name,
                'region': site_config.get('region', ''),
                'error': str(e),
                'prices': {}
            })
    
    # 価格修正マッピングを適用
    if price_corrections:
        logger.info("価格修正マッピングを適用中...")
        results = apply_price_corrections(results, price_corrections)
    
    # 結果を保存
    save_results(results, output_format='json')
    save_results(results, output_format='csv')
    save_results(results, output_format='excel')
    
    # サマリー表示
    success_count = sum(1 for r in results if r.get('prices'))
    total_prices = sum(len(r.get('prices', {})) for r in results)
    
    logger.info(f"\n{'='*60}")
    logger.info(f"スクレイピング完了:")
    logger.info(f"  成功: {success_count}/{len(sites)} 社")
    logger.info(f"  失敗: {len(sites) - success_count}/{len(sites)} 社")
    logger.info(f"  取得価格情報総数: {total_prices} 件")
    logger.info(f"{'='*60}")


if __name__ == '__main__':
    main()

