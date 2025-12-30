#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
汎用的な表形式シートへの価格記入システム
設定ファイル（config/output_tables.yaml）で指定した複数のExcelファイル・シートに
自動的に価格を記入します
"""

import yaml
import logging
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from scrapers import Category1Scraper, Category2Scraper

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scrape_log_v2.txt', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 実装済み18社のリスト（正規化された企業名）
IMPLEMENTED_COMPANIES = {
    '眞田鋼業株式会社',
    '有限会社金田商事',
    '木村金属（大阪）',
    '明鑫貿易株式会社',
    '東起産業（株）',
    '土金（大阪）',
    '大畑商事（千葉・大阪）',
    '千福商会（大阪）',
    '鴻祥貿易株式会社',
    '株式会社鳳山',
    '株式会社 春日商会　富山支店',
    '株式会社 春日商会　滋賀支店',
    '株式会社 春日商会　一宮本社',
    '安城貿易（愛知）',
    '東北キング',
    '株式会社八木',
    '有限会社　八尾アルミセンター',
    '株式会社 ヒラノヤ',
    '鴻陽産業株式会社 岐阜工場',
    '株式会社 大垣金属',
    '高橋商事株式会社',
}

# 材料名のマッピング（取得した材料名 → シートの列名）
MATERIAL_MAPPING = {
    'ピカ銅': 'ピカ銅',
    'ピカ線': 'ピカ銅',
    'ピカドウ': 'ピカ銅',
    '1号銅': 'ピカ銅',
    '一号銅': 'ピカ銅',
    '特一号銅': 'ピカ銅',
    '特1号銅': 'ピカ銅',
    '一号銅線（ピカ線）': 'ピカ銅',
    'ピカ線一号': 'ピカ銅',
    '上銅': 'ピカ銅',
    '上故銅': 'ピカ銅',
    'ピカ線(1号銅線)': 'ピカ銅',
    
    '並銅': '並銅',
    '波銅': '並銅',
    '波道': '並銅',
    'なみどう': '並銅',
    '銅（並）銅管': '並銅',
    '銅(並)銅管': '並銅',
    
    '砲金': '砲金',
    'ほうきん': '砲金',
    'gunmetal': '砲金',
    '青銅': '砲金',
    
    '真鍮': '真鍮',
    'しんちゅう': '真鍮',
    '黄銅': '真鍮',
    'brass': '真鍮',
    '真鍮A': '真鍮',
    '込真鍮': '真鍮',
    '込真鍮（黄銅）': '真鍮',
    
    '雑線80%': '雑線80%',
    '雑電線80%': '雑線80%',
    '電線80%': '雑線80%',
    '銅率80%': '雑線80%',
    '銅80%': '雑線80%',
    '80%線': '雑線80%',
    '一本線80%': '雑線80%',
    '一本線(A)': '雑線80%',
    '上線（80％）': '雑線80%',
    
    '雑線60%': '雑線60%-65%',
    '雑線65%': '雑線60%-65%',
    '雑電線60%': '雑線60%-65%',
    '雑電線65%': '雑線60%-65%',
    '電線60%': '雑線60%-65%',
    '電線65%': '雑線60%-65%',
    '銅率60%': '雑線60%-65%',
    '銅率65%': '雑線60%-65%',
    '銅60%': '雑線60%-65%',
    '銅65%': '雑線60%-65%',
    '60%線': '雑線60%-65%',
    '雑線60-65%': '雑線60%-65%',
    '三本線65%': '雑線60%-65%',
    '三本線(A)': '雑線60%-65%',
    '上線（60％）': '雑線60%-65%',
    
    'VA線': 'VA線',
    'VVF': 'VA線',
    'VVFケーブル': 'VA線',
    'VA線・巻物': 'VA線',
    
    'アルミホイール': 'アルミホイール',
    'ホイール': 'アルミホイール',
    'アルミホイル': 'アルミホイール',
    
    'アルミサッシ': 'アルミサッシ',
    'サッシ': 'アルミサッシ',
    'サッシ新': 'アルミサッシ',
    'サッシA': 'アルミサッシ',
    'サッシB': 'アルミサッシ',
    'アルミサッシA': 'アルミサッシ',
    'アルミサッシ(ビス付き)': 'アルミサッシ',
    
    'アルミ缶バラ': 'アルミ缶　バラ',
    '缶バラ': 'アルミ缶　バラ',
    'アルミ缶': 'アルミ缶　バラ',
    'バラアルミ缶': 'アルミ缶　バラ',
    
    'アルミ缶プレス': 'アルミ缶　プレス',
    '缶プレス': 'アルミ缶　プレス',
    'アルミプレス': 'アルミ缶　プレス',
    'アルミ缶（プレス）': 'アルミ缶　プレス',
    
    'SUS304': 'ステンレス304',
    'ステンレス304': 'ステンレス304',
    '304': 'ステンレス304',
    'ステンレス': 'ステンレス304',
    
    '鉛バッテリー': '鉛バッテリー',
    'バッテリー': '鉛バッテリー',
    '車バッテリー': '鉛バッテリー',
    '鉛': '鉛バッテリー',
}

# 企業名のマッピング（文字化け対応）
COMPANY_NAME_MAPPING = {
    '眞田鋼業株式会社': '眞田鋼業株式会社',
    '明鑫貿易株式会社': '明鑫貿易�式会社',
    '明鑫貿易�式会社': '明鑫貿易株式会社',
    '東起産業（株）': '東起産業��檼',
    '東起産業��檼': '東起産業（株）',
    '鴻祥貿易株式会社': '鴻祥貿易�式会社',
    '鴻祥貿易�式会社': '鴻祥貿易株式会社',
    '安城貿易（愛知）': '安城貿易（�知�',
    '安城貿易（�知�': '安城貿易（愛知）',
    '千福商会（大阪）': '卦�商会（大阪�',
    '卦�商会（大阪�': '千福商会（大阪）',
    '土金（大阪）': '土�߼�大阪�',
    '土�߼�大阪�': '土金（大阪）',
    '大畑商事（千葉・大阪）': '大畑商事（千葉�大阪�',
    '大畑商事（千葉�大阪�': '大畑商事（千葉・大阪）',
    '木村金属（大阪）': '木村��属（大阪�',
    '木村��属（大阪�': '木村金属（大阪）',
    '株式会社 春日商会　富山支店': '株式会社 春日啼 富山支�',
    '株式会社 春日啼 富山支�': '株式会社 春日商会　富山支店',
    '株式会社 春日商会　滋賀支店': '株式会社 春日啼 滋�支�',
    '株式会社 春日啼 滋�支�': '株式会社 春日商会　滋賀支店',
    '株式会社 春日商会　一宮本社': '株式会社 春日啼 �宮本社',
    '株式会社 春日啼 �宮本社': '株式会社 春日商会　一宮本社',
    '株式会社八木': '株式会社八木',
    '有限会社　八尾アルミセンター': '有限会社　八尾アルミセンター',
    '株式会社 ヒラノヤ': '株式会社 ヒラノヤ',
    '株式会社鳳山': '株式会社鳳山',
    '東北キング': '東北キング',
    '有限会社金田商事': '有限会社金田商事',
}

def normalize_company_name(name):
    """企業名を正規化（文字化け対応、重複を避ける）"""
    if not name:
        return ''
    
    name = str(name).strip()
    
    # マッピングを確認
    if name in COMPANY_NAME_MAPPING:
        return COMPANY_NAME_MAPPING[name]
    
    # 部分一致でマッピングを探す
    for key, value in COMPANY_NAME_MAPPING.items():
        if key in name or name in key:
            return value
    
    # 実装済み18社のリストと照合
    for implemented_name in IMPLEMENTED_COMPANIES:
        # 部分一致で確認
        if implemented_name in name or name in implemented_name:
            return implemented_name
    
    return name

def normalize_price(price_str):
    """価格文字列を正規化（数値のみを抽出）"""
    if not price_str:
        return ''
    
    # 数値を抽出
    price_match = re.search(r'(\d{1,4}(?:[,，]\d{3})*(?:\.\d+)?)', str(price_str))
    if price_match:
        price_value = price_match.group(1).replace(',', '').replace('，', '')
        return price_value
    return ''

def load_site_config(config_path: str = 'config/sites.yaml'):
    """サイト設定ファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('sites', [])
    except Exception as e:
        logger.error(f"設定ファイルの読み込みエラー: {str(e)}")
        return []

def filter_implemented_companies(sites):
    """実装済み企業のみをフィルタリング"""
    filtered = []
    seen_companies = set()
    
    for site in sites:
        company_name = site.get('name', '')
        normalized_name = normalize_company_name(company_name)
        
        if normalized_name not in IMPLEMENTED_COMPANIES:
            continue
        
        if normalized_name in seen_companies:
            logger.warning(f"重複をスキップ: {company_name} (正規化後: {normalized_name})")
            continue
        
        seen_companies.add(normalized_name)
        filtered.append(site)
    
    return filtered

def load_target_items_config(config_path: str = 'config/target_items.yaml'):
    """対象アイテム設定ファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('target_items', [])
    except Exception as e:
        logger.warning(f"対象アイテム設定ファイルの読み込みエラー: {str(e)}")
        return []

def load_price_corrections(config_path: str = 'config/price_corrections.yaml'):
    """価格修正マッピングファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('corrections', {})
    except Exception as e:
        logger.warning(f"価格修正マッピングファイルの読み込みエラー: {str(e)}")
        return {}

def apply_price_corrections(results, corrections):
    """価格修正マッピングを適用"""
    corrected_results = []
    
    for result in results:
        company_name = result.get('company_name', '')
        prices = result.get('prices', {}).copy()
        
        if company_name in corrections:
            correction = corrections[company_name]
            
            if 'remove' in correction:
                for material in correction['remove']:
                    if material in prices:
                        del prices[material]
            
            if 'add' in correction:
                for item in correction['add']:
                    prices[item['material']] = item['price']
            
            if 'modify' in correction:
                for item in correction['modify']:
                    old_material = item['material']
                    new_price = item['price']
                    new_material = item.get('material_new', old_material)
                    
                    if old_material in prices:
                        if new_material != old_material:
                            if new_price == new_material or new_price == old_material:
                                prices[new_material] = prices[old_material]
                            else:
                                prices[new_material] = new_price
                            del prices[old_material]
                        else:
                            prices[old_material] = new_price
        
        corrected_result = result.copy()
        corrected_result['prices'] = prices
        corrected_results.append(corrected_result)
    
    return corrected_results

def scrape_implemented_companies():
    """実装済み18社の価格データをスクレイピングで取得"""
    site_configs = load_site_config()
    target_items_config = load_target_items_config()
    price_corrections = load_price_corrections()
    
    sites = filter_implemented_companies(site_configs)
    
    logger.info(f"実装済み企業: {len(sites)}社を対象にスクレイピングを開始します")
    for i, site in enumerate(sites, 1):
        logger.info(f"  {i}. {site.get('name', '不明')}")
    
    company_results = []
    
    for i, site_config in enumerate(sites, 1):
        company_name = site_config.get('name', '不明')
        category = site_config.get('category', 2)
        normalized_name = normalize_company_name(company_name)
        
        logger.info(f"\n[{i}/{len(sites)}] 処理中: {company_name} (正規化後: {normalized_name})")
        
        try:
            if category == 1:
                scraper = Category1Scraper(site_config, delay=2.0)
            elif category == 2:
                scraper = Category2Scraper(site_config, delay=2.0)
            else:
                logger.warning(f"  不明なカテゴリ: {category}")
                continue
            
            result = scraper.scrape(
                filter_target_items=True,
                target_items_config=target_items_config
            )
            
            result['company_name'] = normalized_name
            company_results.append(result)
            
            prices = result.get('prices', {})
            if prices:
                price_count = len(prices)
                logger.info(f"  ✓ {price_count} 件の価格情報を取得")
                for material, price in list(prices.items())[:5]:
                    logger.info(f"    - {material}: {price}")
            else:
                error = result.get('error', '')
                logger.warning(f"  ✗ 価格情報を取得できませんでした: {error}")
        
        except Exception as e:
            logger.error(f"  エラー: {company_name} - {str(e)}")
            company_results.append({
                'scraped_at': datetime.now().isoformat(),
                'url': site_config.get('price_url', ''),
                'company_name': normalized_name,
                'region': site_config.get('region', ''),
                'error': str(e),
                'prices': {}
            })
    
    if price_corrections:
        logger.info("価格修正マッピングを適用中...")
        company_results = apply_price_corrections(company_results, price_corrections)
    
    return company_results

def fill_table_format(excel_file, company_results, target_sheet_name):
    """
    表形式のシートにスクレイピング結果を記入（汎用版）
    
    Args:
        excel_file: Excelファイルのパス
        company_results: スクレイピング結果のリスト
        target_sheet_name: 対象シート名（全角・半角の数字に対応）
    """
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        logger.error(f"エラー: Excelファイルが見つかりません: {excel_file}")
        return False
    except Exception as e:
        logger.error(f"エラー: Excelファイルの読み込みに失敗しました: {excel_file} - {str(e)}")
        return False
    
    # シート名を探す（全角・半角両方に対応）
    actual_sheet_name = None
    for sheet_name in wb.sheetnames:
        # 完全一致または部分一致で探す
        if (target_sheet_name == sheet_name or 
            target_sheet_name in str(sheet_name) or 
            str(sheet_name) in target_sheet_name):
            actual_sheet_name = sheet_name
            break
    
    if not actual_sheet_name:
        logger.error(f"エラー: 「{target_sheet_name}」シートが見つかりません")
        logger.info(f"利用可能なシート: {wb.sheetnames}")
        return False
    
    ws = wb[actual_sheet_name]
    logger.info(f"シート「{actual_sheet_name}」を読み込みました")
    
    # ヘッダー行を取得（1行目、2列目以降）
    header_materials = {}
    for col_idx in range(2, ws.max_column + 1):  # 2列目から（1列目は企業名）
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            header_materials[str(cell.value).strip()] = col_idx
    
    logger.info(f"ヘッダー材料: {list(header_materials.keys())}")
    
    # 既存の企業名のリストを作成（2行目以降）
    existing_companies = {}
    for row_idx in range(2, ws.max_row + 1):
        company_name_cell = ws.cell(row=row_idx, column=1)
        company_name = str(company_name_cell.value).strip() if company_name_cell.value else ''
        if company_name:
            normalized = normalize_company_name(company_name)
            existing_companies[normalized] = row_idx
    
    logger.info(f"既存の企業: {len(existing_companies)}社")
    
    # スクレイピングで取得した企業の価格を記入
    processed_companies = set()
    filled_count = 0
    next_row = ws.max_row + 1
    
    for result in company_results:
        company_name = result.get('company_name', '')
        normalized_name = normalize_company_name(company_name)
        prices = result.get('prices', {})
        
        if not prices:
            logger.warning(f"  {company_name}: 価格データがありません")
            continue
        
        if normalized_name in processed_companies:
            logger.warning(f"重複をスキップ: {company_name} (正規化後: {normalized_name})")
            continue
        
        processed_companies.add(normalized_name)
        
        # 既存の企業か確認
        row_idx = None
        if normalized_name in existing_companies:
            row_idx = existing_companies[normalized_name]
            logger.info(f"  {company_name}: 既存の行{row_idx}に記入")
        else:
            # 新しい行に企業名を追加
            ws.cell(row=next_row, column=1, value=normalized_name)
            row_idx = next_row
            existing_companies[normalized_name] = row_idx
            logger.info(f"  {company_name}: 新規追加 (行{next_row})")
            next_row += 1
        
        # 各材料の価格を記入
        for material_name, price_value in prices.items():
            # 材料名を正規化
            normalized_material = None
            for key, value in MATERIAL_MAPPING.items():
                if key in material_name or material_name in key:
                    normalized_material = value
                    break
            
            if not normalized_material:
                if material_name in header_materials:
                    normalized_material = material_name
                else:
                    logger.debug(f"    未マッチ: {material_name}")
                    continue
            
            # 列番号を取得（全角スペースの違いを考慮）
            col_idx = None
            if normalized_material in header_materials:
                col_idx = header_materials[normalized_material]
            else:
                normalized_material_alt = normalized_material.replace('　', ' ')
                if normalized_material_alt in header_materials:
                    col_idx = header_materials[normalized_material_alt]
                else:
                    for header_key in header_materials.keys():
                        if (normalized_material.replace(' ', '　') == header_key or 
                            normalized_material == header_key.replace(' ', '　')):
                            col_idx = header_materials[header_key]
                            break
            
            if not col_idx:
                logger.debug(f"    警告: {normalized_material}の列が見つかりません")
                continue
            
            # 価格を正規化
            normalized_price = normalize_price(price_value)
            
            if normalized_price:
                ws.cell(row=row_idx, column=col_idx, value=normalized_price)
                filled_count += 1
                logger.info(f"    {normalized_material}: {normalized_price}円 (行{row_idx}, 列{col_idx})")
    
    # 罫線を追加
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    logger.info("\n罫線を追加中...")
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    # ファイルを保存
    try:
        wb.save(excel_file)
        logger.info(f"\n✓ 「{actual_sheet_name}」シートにスクレイピング結果を記入しました")
        logger.info(f"  ファイル: {excel_file}")
        logger.info(f"  処理した企業数: {len(processed_companies)}社")
        logger.info(f"  記入した価格数: {filled_count}件")
        return True
    except Exception as e:
        logger.error(f"エラー: ファイルの保存に失敗しました: {str(e)}")
        return False

def load_output_tables_config(config_path: str = 'config/output_tables.yaml'):
    """出力先テーブル設定ファイルを読み込む"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            return config.get('output_tables', [])
    except FileNotFoundError:
        logger.warning(f"出力先テーブル設定ファイルが見つかりません: {config_path}")
        logger.info("デフォルト設定を使用します")
        return []
    except Exception as e:
        logger.error(f"出力先テーブル設定ファイルの読み込みエラー: {str(e)}")
        return []

if __name__ == '__main__':
    logger.info("="*80)
    logger.info("汎用的な表形式シートへの価格記入システムを開始します...")
    logger.info("="*80)
    
    # スクレイピングを実行
    company_results = scrape_implemented_companies()
    
    # サマリー表示
    success_count = sum(1 for r in company_results if r.get('prices'))
    total_prices = sum(len(r.get('prices', {})) for r in company_results)
    
    logger.info(f"\n{'='*60}")
    logger.info(f"スクレイピング完了:")
    logger.info(f"  成功: {success_count}/{len(company_results)} 社")
    logger.info(f"  失敗: {len(company_results) - success_count}/{len(company_results)} 社")
    logger.info(f"  取得価格情報総数: {total_prices} 件")
    logger.info(f"{'='*60}")
    
    # 出力先テーブル設定を読み込む
    output_tables = load_output_tables_config()
    
    if not output_tables:
        logger.warning("出力先テーブル設定がありません。config/output_tables.yaml を設定してください。")
    else:
        logger.info(f"\n出力先テーブル設定: {len(output_tables)}件")
        
        # 各出力先に記入
        success_tables = 0
        for i, table_config in enumerate(output_tables, 1):
            excel_file = table_config.get('excel_file', '')
            sheet_name = table_config.get('sheet_name', '')
            description = table_config.get('description', '')
            enabled = table_config.get('enabled', True)
            
            if not enabled:
                logger.info(f"\n[{i}/{len(output_tables)}] スキップ: {excel_file} - {sheet_name} ({description})")
                continue
            
            if not excel_file or not sheet_name:
                logger.warning(f"[{i}/{len(output_tables)}] 設定が不完全です: {table_config}")
                continue
            
            logger.info(f"\n[{i}/{len(output_tables)}] 処理中: {excel_file} - {sheet_name}")
            if description:
                logger.info(f"  説明: {description}")
            
            if fill_table_format(excel_file, company_results, sheet_name):
                success_tables += 1
        
        logger.info(f"\n{'='*60}")
        logger.info(f"表形式シートへの記入完了:")
        logger.info(f"  成功: {success_tables}/{len([t for t in output_tables if t.get('enabled', True)])} シート")
        logger.info(f"{'='*60}")
    
    logger.info("\n完了しました！")

